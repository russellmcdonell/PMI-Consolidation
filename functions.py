'''The common functions'''

# pylint: disable=invalid-name, bare-except, line-too-long, too-many-lines, unspecified-encoding

import os
import sys
import logging
import csv
import re
import datetime
from configparser import ConfigParser as ConfParser
from configparser import MissingSectionHeaderError, NoSectionError, NoOptionError, ParsingError
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from jellyfish import soundex, nysiis, metaphone, levenshtein_distance, jaro_winkler_similarity
import data as d

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0        # successful termination
EX_WARN = 1        # non-fatal termination with warnings

EX_USAGE = 64        # command line usage error
EX_DATAERR = 65        # data format error
EX_NOINPUT = 66        # cannot open input
EX_NOUSER = 67        # addressee unknown
EX_NOHOST = 68        # host name unknown
EX_UNAVAILABLE = 69    # service unavailable
EX_SOFTWARE = 70    # internal software error
EX_OSERR = 71        # system error (e.g., can't fork)
EX_OSFILE = 72        # critical OS file missing
EX_CANTCREAT = 73    # can't create (user) output file
EX_IOERR = 74        # input/output error
EX_TEMPFAIL = 75    # temp failure; user is invited to retry
EX_PROTOCOL = 76    # remote error in protocol
EX_NOPERM = 77        # permission denied
EX_CONFIG = 78        # configuration error



def getMasterConfig(extractDir):
    '''Read in the master configuration data.
(masterDir/master.cfg or masterDir/extractDir/extract.cfg)
After reading in the configuration data, check that it defines a valid configuation - mandatory things are defined and references to things do reference defined things.'''

    if not os.path.exists(f'./{d.masterDir}') :    # Check masterDir exists
        logging.fatal('master directory ./%s not found', d.masterDir)
        sys.exit(EX_NOINPUT)
    if extractDir:
        if not os.path.exists(f'./{d.masterDir}/{d.masterExtractDir}') :    # Check masterDir/masterExtractDir exists
            logging.fatal('master directory ./%s/%s not found', d.masterDir, d.masterExtractDir)
            sys.exit(EX_NOINPUT)
        if not os.path.exists(f'./{d.masterDir}/{d.masterExtractDir}/extract.cfg') :    # check masterDir/masterExtractDir/extract.cfg exists
            return
    else:
        if not os.path.exists(f'./{d.masterDir}/master.cfg') :    # check masterDir/master.cfg exists
            logging.fatal('./%s/master.cfg not found', d.masterDir)
            sys.exit(EX_NOINPUT)

    # Set up a configuation parser and try reading and checking master.cfg
    config = ConfParser()
    config.optionxform = str
    try:
        if extractDir:
            config.read(f'./{d.masterDir}/{d.masterExtractDir}/extract.cfg')
        else:
            config.read(f'./{d.masterDir}/master.cfg')

        # Check Section [masterFile]
        if extractDir:
            d.masterFileName = './' + d.masterDir + '/' + d.masterExtractDir + '/' + config.get('masterFile', 'masterFileName')
        else:
            d.masterFileName = './' + d.masterDir + '/' + config.get('masterFile', 'masterFileName')
        if config.has_option('masterFile', 'masterFieldCount') or not extractDir:
            d.masterFieldCount = config.getint('masterFile', 'masterFieldCount')

        # Check Section [masterSaveColumns]
        if config.has_section('masterSaveColumns') or not extractDir:
            cols = config.get('masterSaveColumns', 'master save columns')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.masterSaveColumns=row
                break
            for i, col in enumerate(d.masterSaveColumns):
                d.masterSaveColumns[i] = int(col)
            cols = config.get('masterSaveColumns', 'master save titles')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.masterSaveTitles=row
                break
            for i, title in enumerate(d.masterSaveTitles):
                d.masterIs[title] = i
                d.masterHas[title] = title


        # Check Section [masterNames]
        if config.has_section('masterNames') or not extractDir:
            d.masterShortName = config.get('masterNames', 'master short name')
            d.masterLongName = config.get('masterNames', 'master long name')

        # Check Section [masterIDs]
        if config.has_section('masterIDs') or not extractDir:
            d.masterPIDname = config.get('masterIDs', 'masterPIDname')
            d.masterURname = config.get('masterIDs', 'masterURname')

        # Check Section [masterNextUR]
        if config.has_section('masterNextUR') or not extractDir:
            d.masterNextUR = config.get('masterNextUR', 'next master UR')


        # Check Section [masterConcepts]
        if config.has_section('masterConcepts') or not extractDir:
            masterConcepts = config.items('masterConcepts')
            for item in (masterConcepts):
                if item[1] and (item[1] != 'None'):
                    d.masterHas[item[0]] = item[1]

        # Check Section [masterMerged]
        if config.has_section('masterMerged') or not extractDir:
            d.masterLinks['mergedLink'] = config.get('masterMerged', 'master merged link')
            if d.masterLinks['mergedLink'] == 'None':
                d.masterLinks['mergedLink'] = None
            if d.masterLinks['mergedLink'] is not None:
                d.masterLinks['mergedIs'] = config.get('masterMerged', 'master merged is')

        # Check Section [masterAlias]
        if config.has_section('masterAlias') or not extractDir:
            d.masterLinks['aliasLink'] = config.get('masterAlias', 'master alias link')
            if d.masterLinks['aliasLink'] == 'None':
                d.masterLinks['aliasLink'] = None

        # Check Section [masterReporting]
        if config.has_section('masterReporting'):
            cols = config.get('masterReporting', 'master reporting columns')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.masterReportingColumns = row
                break

        # Check the Extensive checking Sections
        if d.Extensive:
            if config.has_option('ExtensiveConfidence', 'Check Confidence Value') or not extractDir:
                d.ExtensiveConfidence = config.getfloat('ExtensiveConfidence', 'Check Confidence Value')
            if config.has_section('ExtensiveCore'):
                d.ExtensiveRoutines = {}
                ExtensiveCore = config.items('ExtensiveCore')
                for item in (ExtensiveCore):
                    if item[0] not in (d.DefinedRoutines):
                        logging.fatal('Extensive checking invoked, but unknowns core data checking routine (%s) specified in ./%s/master.cfg', item[0], d.masterDir)
                        sys.exit(EX_CONFIG)
                    if 'Near' in item[0]:
                        if ',' not in item[1]:
                            logging.fatal('Extensive checking routine "%s" specified specified in ./%s/master.cfg but no parameter specified', item[0], d.masterDir)
                            sys.exit(EX_CONFIG)
                        (weight, param) = item[1].split(',')
                        try:
                            w = float(weight)
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad weight (%s) in ./%s/master.cfg: %s', item[0], weight, d.masterDir, e)
                            sys.exit(EX_CONFIG)
                        try:
                            p = int(param)
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad parameter (%s) in ./%s/master.cfg: %s', item[0], param, d.masterDir, e)
                            sys.exit(EX_CONFIG)
                        if w != 0:
                            d.ExtensiveRoutines[item[0]] = (w, p)
                    else:
                        try:
                            w = float(item[1])
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad weight (%s) in ./%s/master.cfg', item[0], item[1], d.masterDir)
                            sys.exit(EX_CONFIG)
                        if w != 0:
                            if 'Middle' in item[0]:
                                if 'MiddleNames' not in d.masterHas:
                                    logging.fatal('Extensive checking invoked with routine "%s" in ./%s/master.cfg, but "MiddleNames" not found in the [masterConcepts] section', item[0], d.masterDir)
                                    sys.exit(EX_CONFIG)
                                else:
                                    d.useMiddleNames = True
                            d.ExtensiveRoutines[item[0]] = w
            if config.has_section('ExtensiveFields'):
                d.ExtensiveFields = {}
                ExtensiveFields = config.items('ExtensiveFields')
                for item in (ExtensiveFields):
                    if item[0] not in d.masterHas:
                        logging.fatal('Extensive checking invoked for "other colum" (%s) in ./%s/master.cfg, but this column not found in the [masterConcepts] section', item[0], d.masterDir)
                        sys.exit(EX_CONFIG)
                    try:
                        w = float(item[1])
                    except (ValueError) as e:
                        logging.fatal('Extensive checking for "other column" (%s) with bad weight (%s) in ./%s/master.cfg: %s', item[0], item[1], d.masterDir, e)
                        sys.exit(EX_CONFIG)
                    if w != 0:
                        d.ExtensiveFields[item[0]] = w

    except (MissingSectionHeaderError, NoSectionError, NoOptionError, ParsingError) as e:
        logging.fatal('%s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.masterDir, d.masterExtractDir)
        else:
            logging.fatal('./%s/master.cfg configuration error', d.masterDir)
        sys.exit(EX_CONFIG)

    except (NameError, RuntimeError) as e:
        logging.fatal('%s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.masterDir, d.masterExtractDir)
        else:
            logging.fatal('./%s/master.cfg configuration error', d.masterDir)
        sys.exit(EX_CONFIG)

    except (KeyError) as e:
        logging.fatal('KeyError: %s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.masterDir, d.masterExtractDir)
        else:
            logging.fatal('./%s/master.cfg configuration error', d.masterDir)
        sys.exit(EX_CONFIG)

    except SystemExit:
        sys.exit(EX_OSERR)

    except:
        logging.fatal('%s', sys.exc_info()[0])
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.masterDir, d.masterExtractDir)
        else:
            logging.fatal('./%s/master.cfg configuration error', d.masterDir)
        sys.exit(EX_CONFIG)

    # Check that we have a valid configuration - required things are present
    if 'PID' not in d.masterHas:
        logging.fatal('"PID" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if 'UR' not in d.masterHas:
        logging.fatal('"UR" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if 'FamilyName' not in d.masterHas:
        logging.fatal('"FamilyName" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if 'GivenName' not in d.masterHas:
        logging.fatal('"GivenName" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if 'Birthdate' not in d.masterHas:
        logging.fatal('"Birthdate" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if 'Sex' not in d.masterHas:
        logging.fatal('"Sex" not defined in either section [masterConcepts] or section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if not d.masterFileName:
        logging.fatal('"masterFileName" not defined in section [masterFile]')
        sys.exit(EX_CONFIG)

    if not d.masterFieldCount:
        logging.fatal('"masterFieldCount" not defined in section [masterFile]')
        sys.exit(EX_CONFIG)

    if not d.masterSaveColumns:
        logging.fatal('"masterSaveColumns" not defined in section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    if not d.masterSaveTitles:
        logging.fatal('"masterSaveTitles" not defined in section [masterSaveColumns]')
        sys.exit(EX_CONFIG)

    # Check that the master configuration information makes logical sense
    for i, col in enumerate(d.masterSaveColumns):
        if col >= d.masterFieldCount:
            logging.fatal('masterSaveColumns[%d] is set to %d, but the maximumn number of columns is %d', i, col, d.masterFieldCount)
            sys.exit(EX_CONFIG)

    if len(d.masterSaveTitles) > len(d.masterSaveColumns):
        logging.fatal('Configuration error: more save titles than save columns')
        sys.exit(EX_CONFIG)

    if len(d.masterSaveTitles) < len(d.masterSaveColumns):
        logging.fatal('Configuration error: more save columns than save titles')
        sys.exit(EX_CONFIG)

    for concept, title in d.masterHas.items():
        if title not in d.masterIs:
            logging.fatal('Concept "%s" defined in section [masterConcepts] as column "%s", but column "%s" is not defined in save titles in section [masterSaveColumns]', concept, title, title)
            sys.exit(EX_CONFIG)

    if 'Merged' in d.masterHas:
        if 'mergedLink' not in d.masterLinks:
            logging.fatal('Concept "Merged" defined in section [masterConcepts] as "%s", but "mergedLink" is not defined in section [masterLinks]', d.masterHas['Merged'])
            sys.exit(EX_CONFIG)
        if d.masterLinks['mergedLink'] is not None:
            if d.masterLinks['mergedLink'] not in d.masterIs:
                logging.fatal('Concept "mergedLink" defined in section [masterMerged] as "%s", but "%s" is not defined in save titles in section [masterSaveColumns]', d.masterLinks['mergedLink'], d.masterLinks['mergedLink'])
                sys.exit(EX_CONFIG)
            if 'mergedIs' not in d.masterLinks:
                logging.fatal('Concept "Merged" defined in section [masterConcepts] as "%s", but "mergedLink" is not defined in section [masterMergd]', d.masterHas['Merged'])
                sys.exit(EX_CONFIG)

    if 'Alias' in d.masterHas:
        if 'aliasLink' not in d.masterLinks:
            logging.fatal('Concept "Alias" defined in section [masterConcepts] as "%s", but "aliasLink" is not defined in section [masterAlias]', d.masterHas['Alias'])
            sys.exit(EX_CONFIG)
        if d.masterLinks['aliasLink'] is not None:
            if d.masterLinks['aliasLink'] not in d.masterIs:
                logging.fatal('Concept "aliasLink" defined in section [masterAlias] as "%s", but "%s" is not defined in "master save titles" in section [masterSaveColumns]', d.masterLinks['aliasLink'], d.masterLinks['aliasLink'])
                sys.exit(EX_CONFIG)

    for col in d.masterReportingColumns:
        if col not in d.masterHas:
            logging.fatal('Column "%s" defined in section [masterReporting] as "%s", but "%s" is not defined in section [masterConcepts] or section [masterSaveColumns]', col, col, col)
            sys.exit(EX_CONFIG)

    return



def getSecondaryConfig(extractDir):
    '''Read in the secondary configuration data.
(secondaryDir/secondary.cfg [not extractDir] or secondaryDir/secondaryExtractDir/extract.cfg [extractDir])
Some things can only be in secondary.cfg
After reading in the configuration data, check that it defines a valid configuation - mandatory things are defined and references to things do reference defined things.'''

    if not os.path.exists(f'./{d.secondaryDir}') :    # Check secondaryDir exists
        logging.fatal('secondary directory ./%s not found', d.secondaryDir)
        sys.exit(EX_NOINPUT)
    if extractDir:
        if not os.path.exists(f'./{d.secondaryDir}/{d.secondaryExtractDir}') :    # Check secondaryDir/secondaryExtractDir exists
            logging.fatal('secondary directory ./%s/%s not found', d.secondaryDir, d.secondaryExtractDir)
            sys.exit(EX_NOINPUT)
        if not os.path.exists(f'./{d.secondaryDir}/{d.secondaryExtractDir}/extract.cfg') :    # check secondaryDir/secondaryExtractDir/extract.cfg exists
            return
    else:
        if not os.path.exists(f'./{d.secondaryDir}/secondary.cfg') :    # check secondaryDir/secondary.cfg exists
            logging.fatal('./%s/secondary.cfg not found', d.secondaryDir)
            sys.exit(EX_NOINPUT)

    # Set up a configuation parser and try reading and checking secondary.cfg
    config = ConfParser()
    config.optionxform = str
    try:
        if extractDir:
            config.read(f'./{d.secondaryDir}/{d.secondaryExtractDir}/extract.cfg')
        else:
            config.read(f'./{d.secondaryDir}/secondary.cfg')

        # Check Section [secondaryFile]
        if extractDir:
            d.secondaryFileName = './' + d.secondaryDir + '/' + d.secondaryExtractDir + '/' + config.get('secondaryFile', 'secondaryFileName')
        else:
            d.secondaryFileName = './' + d.secondaryDir + '/' + config.get('secondaryFile', 'secondaryFileName')
        if config.has_option('secondaryFile', 'secondaryFieldCount') or not extractDir:
            d.secondaryFieldCount = config.getint('secondaryFile', 'secondaryFieldCount')

        # Check Section [secondarySaveColumns]
        if config.has_section('secondarySaveColumns') or not extractDir:
            cols = config.get('secondarySaveColumns', 'secondary save columns')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.secondarySaveColumns=row
                break
            for i, col in enumerate(d.secondarySaveColumns):
                d.secondarySaveColumns[i] = int(col)
            cols = config.get('secondarySaveColumns', 'secondary save titles')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.secondarySaveTitles=row
                break
            for i, title in enumerate(d.secondarySaveTitles):
                d.secondaryIs[title] = i
                d.secondaryHas[title] = title


        # Check Section [secondaryNames]
        if config.has_section('secondaryNames') or not extractDir:
            d.secondaryShortName = config.get('secondaryNames', 'secondary short name')
            d.secondaryLongName = config.get('secondaryNames', 'secondary long name')

        # Check Section [secondaryIDs]
        if config.has_section('secondaryIDs') or not extractDir:
            d.secondaryPIDname = config.get('secondaryIDs', 'secondaryPIDname')
            d.secondaryURname = config.get('secondaryIDs', 'secondaryURname')
            d.secondaryAltURname = config.get('secondaryIDs', 'secondaryAltURname')

        # Check Section [secondaryNextUR]
        if config.has_section('secondaryNextUR') or not extractDir:
            d.secondaryNextUR = config.get('secondaryNextUR', 'next secondary UR')


        # Check Section [secondaryConcepts]
        if config.has_section('secondaryConcepts') or not extractDir:
            secondaryConcepts = config.items('secondaryConcepts')
            for item in (secondaryConcepts):
                if item[1] and (item[1] != 'None'):
                    d.secondaryHas[item[0]] = item[1]

        # Check Section [secondaryMerged]
        if config.has_section('secondaryMerged') or not extractDir:
            d.secondaryLinks['mergedLink'] = config.get('secondaryMerged', 'secondary merged link')
            if d.secondaryLinks['mergedLink'] == 'None':
                d.secondaryLinks['mergedLink'] = None
            if d.secondaryLinks['mergedLink'] is not None:
                d.secondaryLinks['mergedIs'] = config.get('secondaryMerged', 'secondary merged is')

        # Check Section [secondaryAlias]
        if config.has_section('secondaryAlias') or not extractDir:
            d.secondaryLinks['aliasLink'] = config.get('secondaryAlias', 'secondary alias link')
            if d.secondaryLinks['aliasLink'] == 'None':
                d.secondaryLinks['aliasLink'] = None

        # Check Section [secondaryReporting]
        if config.has_section('secondaryReporting'):
            cols = config.get('secondaryReporting', 'secondary reporting columns')
            dialect = csv.Sniffer().sniff(cols)
            dialect.skipinitialspace = True
            for row in csv.reader([cols], dialect):
                d.secondaryReportingColumns = row
                break

        # Check the Extensive checking Sections
        if d.Extensive:
            if config.has_option('ExtensiveConfidence', 'Check Confidence Value'):
                d.ExtensiveConfidence = config.getfloat('ExtensiveConfidence', 'Check Confidence Value')
            if config.has_section('ExtensiveCore'):
                d.ExtensiveRoutines = {}
                ExtensiveCore = config.items('ExtensiveCore')
                for item in (ExtensiveCore):
                    if item[0] not in (d.DefinedRoutines):
                        logging.fatal('Extensive checking invoked, but unknowns core data checking routine (%s) specified in ./%s/secondary.cfg', item[0], d.secondaryDir)
                        sys.exit(EX_CONFIG)
                    if 'Near' in item[0]:
                        if ',' not in item[1]:
                            logging.fatal('Extensive checking routine "%s" specified specified in ./%s/secondary.cfg but no parameter specified', item[0], d.secondaryDir)
                            sys.exit(EX_CONFIG)
                        (weight, param) = item[1].split(',')
                        try:
                            w = float(weight)
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad weight (%s) in ./%s/secondary.cfg: %s', item[0], weight, d.secondaryDir, e)
                            sys.exit(EX_CONFIG)
                        try:
                            p = int(param)
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad parameter (%s) in ./%s/secondary.cfg: %s', item[0], param, d.secondaryDir, e)
                            sys.exit(EX_CONFIG)
                        if w != 0:
                            if 'Middle' in item[0]:
                                if 'MiddleNames' not in d.secondaryHas:
                                    logging.fatal('Extensive checking invoked with routine "%s" in ./%s/secondary.cfg, but "MiddleNames" not found in the [secondaryConcepts] section', item[0], d.secondaryDir)
                                    sys.exit(EX_CONFIG)
                                else:
                                    d.useMiddleNames = True
                            d.ExtensiveRoutines[item[0]] = (w, p)
                    else:
                        try:
                            w = float(item[1])
                        except (ValueError) as e:
                            logging.fatal('Extensive checking routine "%s" with bad weight (%s) in ./%s/secondary.cfg', item[0], item[1], d.secondaryDir)
                            sys.exit(EX_CONFIG)
                        if w != 0:
                            d.ExtensiveRoutines[item[0]] = w
            if config.has_section('ExtensiveFields'):
                d.ExtensiveFields = {}
                ExtensiveFields = config.items('ExtensiveFields')
                for item in (ExtensiveFields):
                    if item[0] not in d.secondaryHas:
                        logging.fatal('Extensive checking invoked for "other colum" (%s) in ./%s/secondary.cfg, but this column not found in the [secondaryConcepts] section', item[0], d.secondaryDir)
                        sys.exit(EX_CONFIG)
                    try:
                        w = float(item[1])
                    except (ValueError) as e:
                        logging.fatal('Extensive checking for "other column" (%s) with bad weight (%s) in ./%s/secondary.cfg: %s', item[0], item[1], d.secondaryDir, e)
                        sys.exit(EX_CONFIG)
                    if w != 0:
                        d.ExtensiveFields[item[0]] = w

    except (MissingSectionHeaderError, NoSectionError, NoOptionError, ParsingError) as e:
        logging.fatal('%s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.secondaryDir, d.secondaryExtractDir)
        else:
            logging.fatal('./%s/secondary.cfg configuration error', d.secondaryDir)
        sys.exit(EX_CONFIG)

    except (NameError, RuntimeError) as e:
        logging.fatal('%s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.secondaryDir, d.secondaryExtractDir)
        else:
            logging.fatal('./%s/secondary.cfg configuration error', d.secondaryDir)
        sys.exit(EX_CONFIG)

    except (KeyError) as e:
        logging.fatal('KeyError: %s', e)
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.secondaryDir, d.secondaryExtractDir)
        else:
            logging.fatal('./%s/secondary.cfg configuration error', d.secondaryDir)
        sys.exit(EX_CONFIG)

    except SystemExit:
        sys.exit(EX_OSERR)

    except:
        logging.fatal('%s', sys.exc_info()[0])
        if extractDir:
            logging.fatal('./%s/%s/extract.cfg configuration error', d.secondaryDir, d.secondaryExtractDir)
        else:
            logging.fatal('./%s/secondary.cfg configuration error', d.secondaryDir)
        sys.exit(EX_CONFIG)

    # Check that we have a valid configuration - required things are present
    if 'PID' not in d.secondaryHas:
        logging.fatal('"PID" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if 'UR' not in d.secondaryHas:
        logging.fatal('"UR" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if 'FamilyName' not in d.secondaryHas:
        logging.fatal('"FamilyName" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if 'GivenName' not in d.secondaryHas:
        logging.fatal('"GivenName" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if 'Birthdate' not in d.secondaryHas:
        logging.fatal('"Birthdate" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if 'Sex' not in d.secondaryHas:
        logging.fatal('"Sex" not defined in either section [secondaryConcepts] or section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if not d.secondaryFileName:
        logging.fatal('"secondaryFileName" not defined in section [secondaryFile]')
        sys.exit(EX_CONFIG)

    if not d.secondaryFieldCount:
        logging.fatal('"secondaryFieldCount" not defined in section [secondaryFile]')
        sys.exit(EX_CONFIG)

    if not d.secondarySaveColumns:
        logging.fatal('"secondarySaveColumns" not defined in section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    if not d.secondarySaveTitles:
        logging.fatal('"secondarySaveTitles" not defined in section [secondarySaveColumns]')
        sys.exit(EX_CONFIG)

    # Check that the secondary configuration information makes logical sense
    for i, col in enumerate(d.secondarySaveColumns):
        if col >= d.secondaryFieldCount:
            logging.fatal('secondarySaveColumns[%d] is set to %d, but the maximumn number of columns is %d', i, col, d.secondaryFieldCount)
            sys.exit(EX_CONFIG)

    if len(d.secondarySaveTitles) > len(d.secondarySaveColumns):
        logging.fatal('Configuration error: more save titles than save columns')
        sys.exit(EX_CONFIG)

    if len(d.secondarySaveTitles) < len(d.secondarySaveColumns):
        logging.fatal('Configuration error: more save columns than save titles')
        sys.exit(EX_CONFIG)

    for concept, title in d.secondaryHas.items():
        if title not in d.secondaryIs:
            logging.fatal('Concept "%s" defined in section [secondaryConcepts] as column "%s", but column "%s" is not defined in save titles in section [secondarySaveColumns]', concept, title, title)
            sys.exit(EX_CONFIG)

    if 'Merged' in d.secondaryHas:
        if 'mergedLink' not in d.secondaryLinks:
            logging.fatal('Concept "Merged" defined in section [secondaryConcepts] as "%s", but "mergedLink" is not defined in section [secondaryLinks]', d.secondaryHas['Merged'])
            sys.exit(EX_CONFIG)
        if d.secondaryLinks['mergedLink'] is not None:
            if d.secondaryLinks['mergedLink'] not in d.secondaryIs:
                logging.fatal('Concept "mergedLink" defined in section [secondaryMerged] as "%s", but "%s" is not defined in save titles in section [secondarySaveColumns]', d.secondaryHas['mergedLink'], d.secondaryHas['mergedLink'])
                sys.exit(EX_CONFIG)
            if 'mergedIs' not in d.secondaryLinks:
                logging.fatal('Concept "Merged" defined in section [secondaryConcepts] as "%s", but "mergedLink" is not defined in section [secondaryMergd]', d.secondaryHas['Merged'])
                sys.exit(EX_CONFIG)

    if 'Alias' in d.secondaryHas:
        if 'aliasLink' not in d.secondaryLinks:
            logging.fatal('Concept "Alias" defined in section [secondaryConcepts] as "%s", but "aliasLink" is not defined in section [secondaryAlias]', d.secondaryHas['Alias'])
            sys.exit(EX_CONFIG)
        if d.secondaryLinks['aliasLink'] is not None:
            if d.secondaryLinks['aliasLink'] not in d.secondaryIs:
                logging.fatal('Concept "aliasLink" defined in section [secondaryAlias] as "%s", but "%s" is not defined in "secondary save titles" in section [secondarySaveColumns]', d.secondaryLinks['aliasLink'], d.secondaryLinks['aliasLink'])
                sys.exit(EX_CONFIG)

    for col in d.secondaryReportingColumns:
        if col not in d.secondaryHas:
            logging.fatal('Column "%s" defined in section [secondaryReporting] as "%s", but "%s" is not defined in section [secondaryConcepts] or section [secondarySaveColumns]', col, col, col)
            sys.exit(EX_CONFIG)

    return


def openErrorFile():
    '''
Open the error log csv file
    '''

    if d.scriptType == 'master':
        if d.masterExtractDir:
            fileName = f'./{d.masterDir}/{d.masterExtractDir}/{d.progName}_Errors.csv'
        else:
            fileName = f'./{d.masterDir}/{d.progName}_Errors.csv'
    else:
        if d.secondaryExtractDir:
            fileName = f'./{d.secondaryDir}/{d.secondaryExtractDir}/{d.progName}_Errors.csv'
        else:
            fileName = f'./{d.secondaryDir}/{d.progName}_Errors.csv'
    try:
        d.fe = open(fileName, 'wt', newline='')
    except:
        logging.fatal('cannot create %s', fileName)
        sys.exit(EX_CANTCREAT)
    d.feCSV = csv.writer(d.fe, dialect='excel')


def openNameCheck(nameType):
    '''
Open the file for name check messages
    '''

    if nameType == 'FamilyName':
        d.workbook['fnc'] = Workbook()
        d.worksheet['fnc'] = d.workbook['fnc'].active
        if d.scriptType == 'master':
            heading = ['Message','RecordNo',d.masterURname,'Original family name','Cleaned Up family name']
        else:
            heading = ['Message','RecordNo',d.secondaryURname,'Original family name','Cleaned Up family name']
        d.worksheet['fnc'].append(heading)
    else:
        d.workbook['gnc'] = Workbook()
        d.worksheet['gnc'] = d.workbook['gnc'].active
        if d.scriptType == 'master':
            heading = ['Message','RecordNo',d.masterURname,'Original given name','Cleaned Up given name']
        else:
            heading = ['Message','RecordNo',d.secondaryURname,'Original given name','Cleaned Up given name']
        d.worksheet['gnc'].append(heading)


def openProbableDuplicatesCheck():
    '''
Open the file for probable duplicate name check messages
    '''

    if d.scriptType == 'master':
        heading = ['Message','Cleaned Up family name','Cleaned Up Given Name','Date of Birth','Sex',d.masterURname]
    else:
        heading = ['Message','Cleaned Up family name','Cleaned Up Given Name','Date of Birth','Sex',d.secondaryURname]
    d.workbook['pdc'] = Workbook()
    d.worksheet['pdc'] = d.workbook['pdc'].active
    d.worksheet['pdc'].append(heading)


def openPossibleDuplicatesCheck():
    '''
Open the file for possible duplicate name check messages
    '''

    d.workbook['pdc'] = Workbook()
    d.worksheet['pdc'] = d.workbook['pdc'].active
    if d.scriptType == 'master':
        heading = ['', d.masterURname, 'Family Name', 'Given Name', 'Birthdate', 'Sex', 'Confidence', 'FN Sound', 'GN Sound', d.masterURname, 'Family Name', 'Given Name', 'Birthdate', 'Sex']
    else:
        heading = ['', d.secondaryURname, 'Family Name', 'Given Name', 'Birthdate', 'Sex', 'Confidence', 'FN Sound', 'GN Sound', d.secondaryURname, 'Family Name', 'Given Name', 'Birthdate', 'Sex']
    for field in (sorted(d.ExtensiveFields.keys())):
        heading.append(field)
    d.worksheet['pdc'].append(heading)


def openReport():
    '''
Open the report file for appending
    '''

    if d.scriptType == 'master':
        if d.masterExtractDir:
            fileName = f'./{d.masterDir}/{d.masterExtractDir}/{d.masterShortName}_{d.progName}_Report.txt'
        else:
            fileName = f'./{d.masterDir}/{d.masterShortName}_{d.progName}_Report.txt'
    else:
        if d.secondaryExtractDir:
            fileName = f'./{d.secondaryDir}/{d.secondaryExtractDir}/{d.secondaryShortName}_{d.progName}_Report.txt'
        else:
            fileName = f'./{d.secondaryDir}/{d.secondaryShortName}_{d.progName}_Report.txt'
    try:
        d.rpt = open(fileName, 'wt', newline='')
    except:
        logging.fatal('cannot create %s', fileName)
        sys.exit(EX_CANTCREAT)


def cleanUpExcel(sheet):
    '''
Make an Excel spreadsheet look nicer by widening the columns
    '''

    for i, column in enumerate(sheet.iter_cols()):
        length = max(len(str(cell.value) if cell.value is not None else '') for cell in column)
        sheet.column_dimensions[get_column_letter(i + 1)].width = int(length * 1.25)


def masterSaveLinks():
    '''
Save 'source' information (which record has which UR, which record has which PID)
for later checking that alias and merge references point to valid records.

Skip aliases - aliases are always "OUT"; i.e. an alias it is an alias of another record.
That other record has to be uniquely identifyable (UR and possibly PID) but not this one.
If an record is both an alias and merged, then merges must be going 'OUT'.
That is, this record was merged into another record, with that other record being the primary.
Again, that other record, and not this one, must be unique, so we can skip records that are aliases whether they are merged or not.
(You shouldn't be able to merge a valid record INTO an alias record; the merge should be to the primary record, not to an alias)
NOTE: multiple aliases can share a UR value.

Check that no other record has the same UR value as the current record.
Save this record number as the record number for this UR value, so we can check merges to UR number have matching records.

If PID values are being used for merge or alias linkages, then no other non-alias record should have this records PID value.
Check that no other record has the same PID value as the current record.
Save this record number as the record number for this PID value, so we can check merges and alias to PID number have matching records.
    '''

    if d.ml.masterIsAlias():
        return

    # Apart from aliases, everything else must have a unique UR number as we use it for linkage between the secondary PMI
    # and the master PMI. If there are duplicates we assume the last one is correct.
    ur = d.mc.masterCleanUR()
    if ur in d.URrec:
        d.URdup += 1
        d.feCSV.writerow([f'Non-unique {d.masterURname} number ({ur}) records', f'{d.URrec[ur]} and {d.masterRecNo}'])
    d.URrec[ur] = d.masterRecNo

    # If there is any linkage based upon PID then PID must be unique for all non-aliases
    pidLinks = False
    if 'Alias' in d.masterLinks:
        if d.masterLinks['Alias'] == 'PID':
            pidLinks = True
    if 'PID' in d.masterLinks:
        if d.masterLinks['Merged'] == 'PID':
            pidLinks = True
    if pidLinks:
        pid = d.mc.masterCleanPID()
        if pid in d.PIDrec:
            d.PIDdup += 1
            d.feCSV.writerow([f'Non-unique {d.masterPIDname} ({pid}) records', f'{d.PIDrec[pid]} and {d.masterRecNo}'])
        else:
            d.PIDrec[pid]  = d.masterRecNo
    return


def masterSetAlias():
    '''
Save the alias information.
This is set temporarily to either a UR number or a patient ID. (the UR number or patient ID that is the primary for this alias)
masterFindAlias() below has to convert this to a record number
    '''

    d.aCount += 1
    d.masterPrimRec[d.masterRecNo] = masterField('Alias')


def masterSetMerged():
    '''
Save the merged information
This is set temporarily to either a UR number or a patient ID
(the UR number or patient ID that has been merge IN to this record or the UR number or patient ID that this record has been merged OUT to)
masterFindMerged() below has to convert this to a record number
    '''

    d.mCount += 1
    if d.masterLinks['mergedIs'] == 'IN':
        d.masterLinkRec[d.masterRecNo] = masterField('Merged')
    else:
        d.masterNewRec[d.masterRecNo] = masterField('Merged')


def masterFindAliases():
    '''
Find 'aliased to' records by pid or UF for record marked as 'is alias'
    '''

    if 'Alias' not in d.masterHas:
        return
    if d.masterLinks['aliasLink'] is None:
        return

    # Find the master for each alias record
    for recNo, thisPID in d.masterPrimRec.items():
        if d.masterLinks['aliasLink'] == 'PID':
            pid = thisPID
            if pid in d.PIDrec:
                d.masterPrimRec[recNo] = d.PIDrec[pid]
            else:
                d.masterPrimRec[recNo] = None
        else:
            ur = d.masterPrimRec[recNo]
            if ur in d.URrec:
                d.masterPrimRec[recNo] = d.URrec[ur]
            else:
                d.masterPrimRec[recNo] = None
    return


def masterFindMerged():
    '''
Find 'merged to' or 'merged from' records by pid or UR record number for record marked as 'merged'
    '''

    if 'Merged' not in d.masterHas:
        return
    if d.masterLinks['mergedLink'] is None:
        return

    # Find the master for each merged record
    if d.masterLinks['mergedIs'] == 'IN':
        for recNo, thisUR in d.masterLinkRec.items():
            # There can only one record merged into this one.  And it must be unique. Find it and make this record it's the new master record
            #  If it cana't be found it doesn't matter as it cannot cause find or match failures
            if d.masterLinks['mergedLink'] == 'UR':
                ur = thisUR
                if ur in d.URrec:
                    d.masterNewRec[d.URrec[ur]] = recNo
            else:
                pid = d.masterLinkRec[recNo]
                if pid in d.PIDrec:
                    d.masterNewRec[d.PIDrec[pid]] = recNo
    else:
        for recNo, thisUR in d.masterNewRec.items():
            # This record has been merged to another which  must be unique. Find it and make it the new master record this record
            if d.masterLinks['mergedLink'] == 'UR':
                ur = thisUR
                if ur in d.URrec:
                    d.masterNewRec[recNo] = d.URrec[ur]
                else:
                    d.masterNewRec[recNo] = None
            else:
                pid = d.masterNewRec[recNo]
                if pid in d.PIDrec:
                    d.masterNewRec[recNo] = d.PIDrec[pid]
                else:
                    d.masterNewRec[recNo] = None
    return


def masterIntUR():
    '''
Get the UR value from the current cleaned up master PMI record and return an integer version of the UR number as a string of digits
    '''

    ur = d.mc.masterCleanUR()
    intUR = ur
    intUR = re.sub('[^0-9-]', '', intUR)
    if len(intUR) > 1:          # Remove any non-leading minus signs
        intUR = intUR[0:1] + re.sub('-', '', intUR[1:])

    # A UR number where all the digits are '0' is not a blank!
    if not re.match('^0*$', intUR):
        intUR = re.sub('^0*', '', intUR)    # Otherwise stip of leading zeros
    return intUR


def masterIntPID ():
    '''
Get the PID value from the current cleaned up master PMI record and return an integer version of the PID number as a string of digits
    '''

    pid = d.mc.masterCleanPID()
    intPID = pid
    intPID = re.sub('[^0-9-]', '', intPID)
    if len(intPID) > 1:         # Remove any non-leading minus signs
        intPID = intPID[0:1] + re.sub('-', '', intPID[1:])

    # A PID number where all the digits are '0' is not a blank!
    if not re.match('^0*$', intPID):
        intPID = re.sub('^0*', '', intPID)    # Otherwise stip of leading zeros
    return intPID


def masterField(concept):
    ''' Retrieve a master field value'''

    if concept in d.masterHas:
        return d.csvfields[d.masterIs[d.masterHas[concept]]]
    else:
        print(f'masterField {concept} not found')
        return ''



def secondarySaveLinks():
    '''
Save 'source' information (which record has which UR, which record has which PID)
for later checking that alias and merge references point to valid records.

Skip aliases - aliases are always "OUT"; i.e. an alias it is an alias of another record.
That other record has to be uniquely identifyable (UR and possibly PID) but not this one.
If an record is both an alias and merged, then merges must be going 'OUT'.
That is, this record was merged into another record, with that other record being the primary.
Again, that other record, and not this one, must be unique, so we can skip records that are aliases whether they are merged or not.
(You shouldn't be able to merge a valid record INTO an alias record; the merge should be to the primary record, not to an alias)
NOTE: multiple aliases can share a UR value.

Check that no other record has the same UR value as the current record.
Save this record number as the record number for this UR value, so we can check merges to UR number have matching records.

If PID values are being used for merge or alias linkages, then no other non-alias record should have this records PID value.
Check that no other record has the same PID value as the current record.
Save this record number as the record number for this PID value, so we can check merges and alias to PID number have matching records.
    '''

    if d.sl.secondaryIsAlias():
        return

    # Apart from aliases, everything else must have a unique UR number as we use it for linkage between the secondary PMI
    # and the secondary PMI. If there are duplicates we assume the last one is correct.
    ur = d.sc.secondaryCleanUR()
    if ur in d.URrec:
        d.URdup += 1
        d.feCSV.writerow([f'Non-unique {d.secondaryURname} number ({ur}) records', f'{d.URrec[ur]} and {d.secondaryRecNo}'])
    d.URrec[ur] = d.secondaryRecNo

    altUR = d.sc.secondaryCleanAltUR()        # Get the altUR number
    pid = d.sc.secondaryCleanPID()
    if altUR:
        if pid not in d.notMatched :        # Ignore if we already know that this record's AltUR is wrong
            # Check for duplicates
            if altUR in d.AltURrecNo:
                d.AltURdup += 1
                d.feCSV.writerow([f'Non-unique {d.secondaryAltURname} number ({altUR}) records', f'{d.AltURrecNo[altUR]} and {d.secondaryRecNo}'])
            else:
                d.AltURrecNo[altUR] = d.secondaryRecNo

    # If there is any linkage based upon PID then PID must be unique for all non-aliases
    pidLinks = False
    if 'Alias' in d.secondaryLinks:
        if d.secondaryLinks['Alias'] == 'PID':
            pidLinks = True
    if 'PID' in d.secondaryLinks:
        if d.secondaryLinks['Merged'] == 'PID':
            pidLinks = True
    if pidLinks:
        pid = d.sc.secondaryCleanPID()
        if pid in d.PIDrec:
            d.PIDdup += 1
            d.feCSV.writerow([f'Non-unique {d.secondaryPIDname} ({pid}) records', f'{d.PIDrec[pid]} and {d.secondaryRecNo}'])
        else:
            d.PIDrec[pid]  = d.secondaryRecNo
    return


def secondarySetAlias():
    '''
Save the alias information.
This is set temporarily to either a UR number or a patient ID. (the UR number or patient ID that is the primary for this alias)
secondaryFindAlias() below has to convert this to a record number
    '''

    d.aCount += 1
    d.secondaryPrimRec[d.secondaryRecNo] = secondaryField('Alias')


def secondarySetMerged():
    '''
Save the merged information
This is set temporarily to either a UR number or a patient ID
(the UR number or patient ID that has been merge IN to this record or the UR number or patient ID that this record has been merged OUT to)
secondaryFindMerged() below has to convert this to a record number
    '''

    d.mCount += 1
    if d.secondaryLinks['mergedIs'] == 'IN':
        d.secondaryLinkRec[d.secondaryRecNo] = secondaryField('Merged')
    else:
        d.secondaryNewRec[d.secondaryRecNo] = secondaryField('Merged')


def secondaryFindAliases():
    '''
Find 'aliased to' records by pid or UF for record marked as 'is alias'
    '''

    if 'Alias' not in d.secondaryHas:
        return
    if d.secondaryLinks['aliasLink'] is None:
        return

    # Find the secondary for each alias record
    for recNo, thisPID in d.secondaryPrimRec.items():
        if d.secondaryLinks['aliasLink'] == 'PID':
            pid = thisPID
            if pid in d.PIDrec:
                d.secondaryPrimRec[recNo] = d.PIDrec[pid]
            else:
                d.secondaryPrimRec[recNo] = None
        else:
            ur = d.secondaryPrimRec[recNo]
            if ur in d.URrec:
                d.secondaryPrimRec[recNo] = d.URrec[ur]
            else:
                d.secondaryPrimRec[recNo] = None
    return


def secondaryFindMerged():
    '''
Find 'merged to' or 'merged from' records by pid or UR record number for record marked as 'merged'
    '''

    if 'Merged' not in d.secondaryHas:
        return
    if d.secondaryLinks['mergedLink'] is None:
        return

    # Find the secondary for each merged record
    if d.secondaryLinks['mergedIs'] == 'IN':
        for recNo, thisUR in d.secondaryLinkRec.items():
            # There can only one record merged into this one.  And it must be unique. Find it and make this record it's the new secondary record
            #  If it cana't be found it doesn't matter as it cannot cause find or match failures
            if d.secondaryLinks['mergedLink'] == 'UR':
                ur = thisUR
                if ur in d.URrec:
                    d.secondaryNewRec[d.URrec[ur]] = recNo
            else:
                pid = d.secondaryLinkRec[recNo]
                if pid in d.PIDrec:
                    d.secondaryNewRec[d.PIDrec[pid]] = recNo
    else:
        for recNo, thisUR in d.secondaryNewRec.items():
            # This record has been merged to another which  must be unique. Find it and make it the new secondary record this record
            if d.secondaryLinks['mergedLink'] == 'UR':
                ur = thisUR
                if ur in d.URrec:
                    d.secondaryNewRec[recNo] = d.URrec[ur]
                else:
                    d.secondaryNewRec[recNo] = None
            else:
                pid = d.secondaryNewRec[recNo]
                if pid in d.PIDrec:
                    d.secondaryNewRec[recNo] = d.PIDrec[pid]
                else:
                    d.secondaryNewRec[recNo] = None
    return


def secondaryIntUR():
    '''
Get the UR value from the current cleaned up secondary PMI record and return an integer version of the UR number as a string of digits
    '''

    ur = d.sc.secondaryCleanUR()
    intUR = ur
    intUR = re.sub('[^0-9-]', '', intUR)
    if len(intUR) > 1:      # Remove non-leading minus signs
        intUR = intUR[0:1] + re.sub('-', '', intUR[1:])

    # A UR number where all the digits are '0' is not a blank!
    if not re.match('^0*$', intUR):
        intUR = re.sub('^0*', '', intUR)    # Otherwise stip of leading zeros
    return intUR


def secondaryIntAltUR():
    '''
Get the AltUR value from the current cleaned up secondary PMI record and return an integer version of the AltUR number as a string of digits
    '''

    altUR = d.sc.secondaryCleanAltUR()
    intAltUR = altUR
    intAltUR = re.sub('[^0-9-]', '', intAltUR)
    if len(intAltUR) > 1:       # Remove non-leadging minus signes
        intAltUR = intAltUR[0:1] + re.sub('-', '', intAltUR[1:])

    # A AltUR number where all the digits are '0' is not a blank!
    if not re.match('^0*$', intAltUR):
        intAltUR = re.sub('^0*', '', intAltUR)    # Otherwise stip of leading zeros
    return intAltUR


def secondaryIntPID ():
    '''
Get the PID value from the current cleaned up secondary PMI record and return an integer version of the PID number as a string of digits
    '''

    pid = d.sc.secondaryCleanPID()
    intPID = pid
    intPID = re.sub('[^0-9-]', '', intPID)
    if len(intPID) > 1:     # Remove any non-leading minus signs
        intPID = intPID[0:1] + re.sub('-', '', intPID[1:])

    # A PID number where all the digits are '0' is not a blank!
    if not re.match('^0*$', intPID):
        intPID = re.sub('^0*', '', intPID)    # Otherwise stip of leading zeros
    return intPID


def secondaryField(concept):
    '''Retrieve a secondary concept value'''

    if concept in d.secondaryHas:
        return d.csvfields[d.secondaryIs[d.secondaryHas[concept]]]
    else:
        print(f'secondaryField {concept} not found')
        return ''



def secondaryFieldSave(concept, value):
    '''Save a secondary concept value'''

    if concept in d.secondaryHas:
        d.csvfields[d.secondaryIs[d.secondaryHas[concept]]] = value
    else:
        print(f'secondaryField {concept} not found')
    return


def getNotMatched():
    '''
Read in the Not Matched PIDs [notMatched.xlsx]
These are PIDs in the secondary PMI, who's AltUR has been checked and has been found to be incorrect
    '''

    d.notMatched = {}
    try:
        if d.secondaryExtractDir:
            nm = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.xlsx')
        else:
            nm = load_workbook(f'./{d.secondaryDir}/notMatched.xlsx')
        nms = nm.active
    except:
        return

    heading = True
    for row in nms.iter_rows():
        if heading:
            heading = False
            continue
        if len(row) != 1:
            if d.secondaryExtractDir:
                logging.fatal('Input error in ./%s/%s/notMatched.xlsx - wrong number of fields', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Input error in ./%s/notMatched.xlsx - wrong number of fields', d.secondaryDir)
            sys.exit(EX_DATAERR)
        pid = str(row[0].value)
        if pid in d.matchedUR:
            if d.secondaryExtractDir:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/{d.secondaryExtractDir}/matched.xlsx - ignoring'])
            else:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/matched.xlsx - ignoring'])
            continue
        if pid in d.foundUR:
            if d.secondaryExtractDir:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/{d.secondaryExtractDir}/found.xlsx - ignoring'])
            else:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/found.xlsx - ignoring'])
            continue
        if pid in d.notFound:
            if d.secondaryExtractDir:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/{d.secondaryExtractDir}/notFound.xlsx - ignoring'])
            else:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/notMatched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/notFound.xlsx - ignoring'])
            continue
        d.notMatched[pid] = True
    return


def getMatched():
    '''
Read in the Matched PIDs and altURs [matched.xlsx]
These are PIDs in the secondary PMI, who's AltUR has been checked and has been found to be correct
The AltUR here is both this PID's AltUR and a valid UR in the master PMI
    '''

    d.matchedPID = {}
    d.matchedUR = {}
    try:
        if d.secondaryExtractDir:
            m = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/matched.xlsx')
        else:
            m = load_workbook(f'./{d.secondaryDir}/matched.xlsx')
        ms = m.active
    except:
        return

    heading = True
    for row in ms.iter_rows():
        if heading:
            heading = False
            continue
        if len(row) != 2:
            logging.fatal('Input error in ./%s/matched.xlsx - wrong number of fields', d.secondaryDir)
            sys.exit(EX_DATAERR)
        pid = row[0].value
        ur = row[1].value
        if pid in d.foundUR:
            d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/matched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/found.xlsx - ignoring'])
            continue
        if pid in d.notFound:
            d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/matched.xlsx', f'{d.secondaryPIDname} ({pid}) already defined in ./{d.secondaryDir}/notFound.xlsx - ignoring'])
            continue
        if pid in d.matchedUR:
            d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/matched.xlsx', 'Duplicate matched records', f'{pid} = {d.matchedUR[pid]} AND {pid} = {ur} - ignoring both'])
            thisUR = d.matchedUR[pid]
            pids = d.matchedPID[thisUR].split('~')
            otherPIDs = ''
            for thisPID in pids:
                if thisPID != pid:
                    if otherPIDs != '':
                        otherPIDs += '~'
                    otherPIDs += thisPID
            if otherPIDs != '':
                d.matchedPID[thisUR] = otherPIDs
            else:
                del d.matchedPID[thisUR]
            continue
        if ur in d.matchedPID:
            d.feCSV.writerow(['WARNING - CREATING DUPLICATE MATCHES', 'Duplicate matched matches', f'{d.secondaryPIDname}({d.matchedPID[ur]}) = {d.secondaryAltURname}({ur}) and {d.secondaryPIDname}({pid}) = {d.secondaryAltURname}({ur})'])
            d.matchedPID[ur] += '~' + pid
        else:
            d.matchedPID[ur] = pid

        d.matchedUR[pid] = ur




def getFound():
    '''
Read in the Found PID and URs [found.xlsx]
These are PIDs in the secondary PMI with their known, matching UR in the master PMI
If this PID had an AltUR, then this is what that AltUR would be
    '''

    d.foundUR = {}
    try:
        if d.secondaryExtractDir:
            f = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/found.xlsx')
        else:
            f = load_workbook(f'./{d.secondaryDir}/found.xlsx')
        fs = f.active
    except:
        return

    heading = True
    for row in fs.iter_rows():
        if heading:
            heading = False
            continue
        if len(row) != 2:
            if d.secondaryExtractDir:
                logging.fatal('Input error in ./%s/%s/found.xlsx - wrong number of fields', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Input error in ./%s/found.xlsx - wrong number of fields', d.secondaryDir)
            sys.exit(EX_DATAERR)
        pid = str(row[0].value)
        ur = str(row[1].value)
        if pid in d.foundUR:
            if d.secondaryExtractDir:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/{d.secondaryExtractDir}/found.xlsx', 'Duplicate found records', f'{d.secondaryPIDname}({pid}) = {d.masterURname}({d.foundUR[pid]}) AND {d.secondaryPIDname}({pid}) = {d.masterURname}({ur}) - ignoring both'])
            else:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/found.xlsx', 'Duplicate found records', f'{d.secondaryPIDname}({pid}) = {d.masterURname}({d.foundUR[pid]}) AND {d.secondaryPIDname}({pid}) = {d.masterURname}({ur}) - ignoring both'])
            thisUR = d.foundUR[pid]
            pids = d.foundPID[thisUR].split('~')
            otherPIDs = ''
            for thisPID in pids:
                if thisPID != pid:
                    if otherPIDs != '':
                        otherPIDs += '~'
                    otherPIDs += thisPID
            if otherPIDs != '':
                d.foundPID[thisUR] = otherPIDs
            else:
                del d.foundPID[thisUR]
            continue
        if ur in d.foundPID:
            d.feCSV.writerow(['WARNING - CREATING DUPLICATES', f'Duplicate secondary PMI records have the same {d.secondaryAltURname} - {d.secondaryPIDname}({d.foundPID[ur]}) = {d.secondaryAltURname}({ur}) and {d.secondaryPIDname}({pid}) = {d.secondaryAltURname}({ur})'])
            d.foundPID[ur] += '~' + pid
        else:
            d.foundPID[ur] = pid

        d.foundUR[pid] = ur
    return


def getNotFound():
    '''
Read in the Not Found PIDs [notFound.xlsx]
These are PIDs in the secondary PMI, which have been checked and it is known that the patient does not exist in the master PMI
    '''

    d.notFound = {}
    try:
        if d.secondaryExtractDir:
            n = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/notFound.xlsx')
        else:
            n = load_workbook(f'./{d.secondaryDir}/notFound.xlsx')
        ns = n.active
    except:
        return

    heading = True
    for row in ns.iter_rows():
        if heading:
            heading = False
            continue
        if len(row) != 1:
            if d.secondaryExtractDir:
                logging.fatal('Input error in ./%s/%s/notFound.xlsx - wrong number of fields', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Input error in ./%s/notFound.xlsx - wrong number of fields', d.secondaryDir)
            sys.exit(EX_DATAERR)
        pid = str(row[0].value)
        if pid in d.foundUR:
            if d.secondaryExtractDir:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/{d.secondaryExtractDir}/notFound.xlsx', f'pid ({pid}) already defined in ./{d.secondaryDir}/{d.secondaryExtractDir}/found.csv'])
            else:
                d.feCSV.writerow([f'ERROR in ./{d.secondaryDir}/notFound.xlsx', f'pid ({pid}) already defined in ./{d.secondaryDir}/found.csv'])
            continue
        d.notFound[pid] = True
    return


def masterSaveDetails ():
    '''
Save the master reporting information
    '''

    d.masterDetails[d.masterRecNo] = {}
    d.masterDetails[d.masterRecNo]['PID'] = masterField('PID')
    d.masterDetails[d.masterRecNo]['UR'] = masterField('UR')
    for col in d.masterReportingColumns:
        if col in d.masterHas:
            d.masterDetails[d.masterRecNo][col] = masterField(col)


def Soundex(name):
    '''
Compute the Soundex value for a name
    '''

    return soundex(name)



def Sounds(part1, part2):
    '''
Compute the NSYIIS, double metaphones and Soundex codes for two names
    '''

    codes = ''
    codes += nysiis(part1)
    codes += '~'
    codes += metaphone(part1)
    codes += '~'
    codes += soundex(part1)
    codes += '~'
    codes += nysiis(part2)
    codes += '~'
    codes += metaphone(part2)
    codes += '~'
    codes += soundex(part2)
    return codes


def SoundCheck(name1, name1ny, name1dm, name1sx, name2, name2ny, name2dm, name2sx):
    '''
Compute the confidence that two names are the same
    '''

    if name1 == name2 :        # Don't bother with computations if they are equal
        return 100.0

    if name1ny == name2ny :        # nysiis is the strictest - so if matches then it's a match
        return 100.0

    l1 = len(name1)
    l2 = len(name2)
    lmax = max(l1, l2)

    # Weight Double Metaphone by Levenshtein distance
    metaConfidence = (1.0 if name1dm == name2dm else 0.0) * (1.0 - float(levenshtein_distance(name1, name2)) / (float(lmax) * 2.0))

    # Weight Soundex by Jaro-Winkler
    soundexConfidence = (1.0 if name1sx == name2sx else 0.0) * jaro_winkler_similarity(name1, name2)

    return max(metaConfidence, soundexConfidence) * 100.0


def FamilyNameCheck(family1, family2, weight):
    '''
Compute the Family Name confidence
    '''
    logging.debug('FamliyNameCheck:%s:%s', family1, family2)

    if (family1 == '') or (family2 == ''):
        return (0.0, 0.0)

    if family1 == family2:
        return (100.0, weight)
    else:
        return (0.0, weight)


def FamilyNameSoundCheck(family1, family1ny, family1dm, family1sx, family2, family2ny, family2dm, family2sx, weight):
    '''
Compute the Family Name sound confidence
    '''
    logging.debug('FamliyNameSoundCheck:%s:%s', family1, family2)

    if (family1 == '') or (family2 == ''):
        return (0.0, 0.0)

    if family1 == family2 :            # Irrelevant if family names are identical
        return (0.0, 0.0)

    return (SoundCheck(family1, family1ny, family1dm, family1sx, family2, family2ny, family2dm, family2sx), weight)



def GivenNameCheck(given1, given2, weight):
    '''
Compute the Given Name confidence
    '''
    logging.debug('GivenNameCheck:%s:%s', given1, given2)

    if (given1 == '') or (given2 == ''):
        return (0.0, 0.0)

    if given1 == given2:
        return (100.0, weight)
    else:
        return (0.0, weight)



def GivenNameSoundCheck(given1, given1ny, given1dm, given1sx, given2, given2ny, given2dm, given2sx, weight):
    '''
Compute the Given Name sound confidence
    '''
    logging.debug('GivenNameSoundCheck:%s:%s', given1, given2)

    if (given1 == '') or (given2 == ''):
        return (0.0, 0.0)

    if given1 == given2 :            # Irrelevant if given names are identical
        return (0.0, 0.0)

    return (SoundCheck(given1, given1ny, given1dm,  given1sx, given2, given2ny, given2dm,  given2sx), weight)


def MiddleNamesCheck(middle1, middle2, weight):
    '''
Compute the Middle Name confidence
    '''
    logging.debug('MiddleNamesCheck:%s:%s', middle1, middle2)

    if (middle1 == '') or (middle2 == ''):
        return (0.0, 0.0)

    if middle1 == middle2:
        return (100.0, weight)
    else:
        return (0.0, weight)


def MiddleNamesInitialCheck(middle1, middle2, weight):
    '''
Compute the Middle Initial confidence
    '''
    logging.debug('MiddleNamesInitialCheck:%s:%s', middle1, middle2)

    if (middle1 == '') or (middle2 == ''):
        return (0.0, 0.0)

    if middle1[0:1] == middle2[0:1]:
        return (100.0, weight)
    else:
        return (0.0, weight)


def SexCheck(sex1, sex2, weight):
    '''
Compute the Sex confidence
    '''
    logging.debug('SexCheck:%s:%s', sex1, sex2)

    if (sex1 == '') or (sex2 == ''):
        return (0.0, 0.0)

    if sex1 == sex2:
        return (100.0, weight)
    else:
        return (0.0, weight)



def BirthdateCheck(birthdate1, birthdate2, weight):
    '''
Compute the Birthdate confidence
    '''
    logging.debug('BirthdateCheck:%s:%s', birthdate1, birthdate2)

    if (birthdate1 == '') or (birthdate2 == ''):
        return (0.0, 0.0)

    if birthdate1 == birthdate2:
        return (100.0, weight)
    else:
        return (0.0, weight)



def Birthdate (dob):
    '''
Convert a string birthdate to a datetime value
    '''

    (year, month, day) = dob.split('-')
    return(datetime.date(int(year), int(month), int(day)))




def BirthdateNearYearCheck(birthdate1, birthdate2, param, weight):
    '''
Check if two birthdates are within 'param' years (i.e. 1981-11-12 === 1984-08-09 if 'param' is 3, but not if 'param' is 2)
    '''
    logging.debug('BirthdateNearYearCheck:%s:%s', birthdate1, birthdate2)

    if (birthdate1 == '') or (birthdate2 == ''):
        return (0.0, 0.0)

    if birthdate1 == birthdate2 :            # Irrelevant if birthdates are identical
        return (0.0, 0.0)

    (year1, month1, day1) = birthdate1.split('-')
    (year2, month2, day2) = birthdate2.split('-')
    year1 = int(year1)
    month1 = int(month1)
    day1 = int(day1)
    year2 = int(year2)
    month2 = int(month2)
    day2 = int(day2)

    if year1 > year2 :                # Check backwards
        yearDiff = year1 - year2
        if yearDiff > param:
            return (0.0, 0.0)
        if month1 > month2:
            yearDiff += 1
        elif month1 == month2:
            if day1 > day2:
                yearDiff += 1
        if yearDiff > param:
            return (0.0, 0.0)
        else:
            return (100.0, weight)
    else:
        yearDiff = year2 - year1
        if yearDiff > param:
            return (0.0, 0.0)
        if month2 > month1:
            yearDiff += 1
        elif month2 == month1:
            if day2 > day1:
                yearDiff += 1
        if yearDiff > param:
            return (0.0, 0.0)
        else:
            return (100.0, weight)



def BirthdateNearMonthCheck(birthdate1, birthdate2, param, weight):
    '''
Check if two birthdates are within 'param' months (i.e. 1981-11-12 === 1981-08-09 if 'param' is 3, but not if 'param' is 2)
    '''
    logging.debug('BirthdateNearMonthCheck')

    if (birthdate1 == '') or (birthdate2 == ''):
        return (0.0, 0.0)

    if birthdate1 == birthdate2 :            # Irrelevant if birthdates are identical
        return (0.0, 0.0)

    (year1, month1, day1) = birthdate1.split('-')
    (year2, month2, day2) = birthdate2.split('-')
    year1 = int(year1)
    month1 = int(month1)
    day1 = int(day1)
    year2 = int(year2)
    month2 = int(month2)
    day2 = int(day2)

    if year1 > year2 :                # Check backwards
        monthDiff = (year1 - year2) * 12 + month1 - month2
        if month1 == month2:
            if day1 > day2:
                monthDiff += 1
        if monthDiff > param:
            return (0.0, 0.0)
        else:
            return (100.0, weight)
    else:
        monthDiff = (year2 - year1) * 12 + month2 - month1
        if month2 == month1:
            if day2 > day1:
                monthDiff += 1
        if monthDiff > param:
            return (0.0, 0.0)
        else:
            return (100.0, weight)


def BirthdateNearDayCheck(date1, date2, param, weight):
    '''
Check if two birthdates are within 'param' days (i.e. 1981-11-12 === 1981-11-09 if 'param' is 3, but not if 'param' is 2)
    '''
    logging.debug('BirthdateNearDayCheck')

    if (date1 == d.futureBirthdate) or (date2 == d.futureBirthdate):
        return (0.0, 0.0)

    if date1 == date2 :            # Irrelevant if birthdates are identical
        return (0.0, 0.0)

    if date1 > date2:
        dateDiff = date1 - date2
    else:
        dateDiff = date2 - date1
    if dateDiff.days < param:
        return (100.0, weight)
    else:
        return (0.0, weight)


def BirthdateYearSwapCheck(birthdate1, birthdate2, weight):
    '''
Check if two dates match if the last two digits of year are swapped (i.e. 1982-11-15 === 1928-11-15)
    '''
    logging.debug('BirthdateYearSwapCheck:%s:%s', birthdate1, birthdate2)

    if (birthdate1 == '') or (birthdate2 == ''):
        return (0.0, 0.0)

    if birthdate1 == birthdate2 :            # Irrelevant if birthdates are identical
        return (0.0, 0.0)

    (year1, month1, day1) = birthdate1.split('-')
    (year2, month2, day2) = birthdate2.split('-')

    if (month1 != month2) or (day1 != day2):
        return (0.0, weight)

    if year1[0:2] != year2[0:2]:
        return (0.0, weight)

    if year1[2:4] ==  year2[3:4] + year2[2:3]:
        return (100.0, weight)
    else:
        return (0.0, weight)



def BirthdateDayMonthSwapCheck(birthdate1, birthdate2, weight):
    '''
Check if two dates match if the month and day are swapped (i.e. 1982-11-12 === 1982-12-11)
    '''
    logging.debug('BirthdateDaySwapCheck:%s:%s', birthdate1, birthdate2)

    if (birthdate1 == '') or (birthdate2 == ''):
        return (0.0, 0.0)

    if birthdate1 == birthdate2 :            # Irrelevant if birthdates are identical
        return (0.0, 0.0)

    (year1, month1, day1) = birthdate1.split('-')
    (year2, month2, day2) = birthdate2.split('-')

    if year1 != year2:
        return (0.0, weight)

    if (month1 == day2) and (month2 == day1):
        return (100.0, weight)
    else:
        return (0.0, 0.0)


def SaveStatus(secondaryRecNo, status, soundFound):
    '''
Assemble a concatinated list of all the Master file records numbers which have the same 'highest' status
    '''

    if secondaryRecNo in d.recStatus:
        if d.recStatus[secondaryRecNo] > status:
            return                            # Lower status - ignore

        if d.recStatus[secondaryRecNo] < status :            # New higher status
            d.foundRec[secondaryRecNo] = ''        # Start new list
            d.foundSound[secondaryRecNo] = ''        # Start new list
            d.recStatus[secondaryRecNo] = status            # Save status
        if d.foundRec[secondaryRecNo] != '' :            # Concatentate if list already started
            d.foundRec[secondaryRecNo] += '~' + str(d.masterRecNo)        # Append to the list
            d.foundSound[secondaryRecNo] += '~' + soundFound        # Append to the list
        else:
            d.foundRec[secondaryRecNo] = str(d.masterRecNo)        # Start the list
            d.foundSound[secondaryRecNo] = soundFound        # Start the list
    else:
        d.recStatus[secondaryRecNo] = status                # Save status
        d.foundRec[secondaryRecNo] = str(d.masterRecNo)            # Start the list
        d.foundSound[secondaryRecNo] = soundFound            # Start the list


def CheckIfFound():
    '''
Check if this secondary record (PID) is in found.xlsx
    '''

    pid = d.sc.secondaryCleanPID()
    if pid in d.foundUR:
        return True
    else:
        return False



def PrintHeading(thisFile, fileType, message1, message2):
    '''
Open a workbook and print the heading
message1 and message2 get prepended to the heading, if presents
fileType defines the number of UR/altUR fields to prepend to the heading
    1 - prepend a heading column of master PMI PID and UR
    2 - prepend heading columns for secondary PID, secondary UR and altUR
    3 - prepend heading columns for both of the above
    4 - prepend heading columns for both of the above and the extensive match confidence levels
    '''

    heading = []
    if message1 != '':
        heading.append(message1)
    if message2 != '':
        heading.append(message2)
    if fileType == 1:
        heading.append(f'{d.masterShortName} {d.masterPIDname}' )
        heading.append(f'{d.masterShortName} {d.masterURname}')
    elif fileType == 2:
        heading.append(f'{d.secondaryShortName} {d.secondaryPIDname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryURname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryAltURname}')
    elif fileType == 3:
        heading.append(f'{d.secondaryShortName} {d.secondaryPIDname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryURname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryAltURname}')
        heading.append(f'{d.masterShortName} {d.masterPIDname}')
        heading.append(f'{d.masterShortName} {d.masterURname}')
    elif fileType == 4:
        heading.append(f'{d.secondaryShortName} {d.secondaryPIDname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryURname}')
        heading.append(f'{d.secondaryShortName} {d.secondaryAltURname}')
        heading.append(f'{d.masterShortName} {d.masterPIDname}')
        heading.append(f'{d.masterShortName} {d.masterURname}')
        heading.append('Confidence')
        heading.append('FN Sound')
        heading.append('GN Sound')

    for col in d.reportingColumns:
        heading.append(col)
    d.workbook[thisFile] = Workbook()
    d.worksheet[thisFile] = d.workbook[thisFile].active
    d.worksheet[thisFile].append(heading)



def PrintSecondary(checked, isChecked, fileType, thisFile, message, confidence):
    '''
Print a secondary PMI record
checked == True if this file is one that gets checked
fileType defines the number of UR/altUR values to prepend to the heading
    1 - prepend a blank, then confidence [here confidence is the auto-allocated Master PMI]
    2 - prepend data columns for secondary PID and altUR (the default)
    3 - prepend data columns for both of the above (master columns will be blank)
    4 - prepend data columns for both of the above and the extensive match confidence levels (master columns will be blank)
    '''

    record = []
    if checked:
        record = [isChecked] + record
    record.append(message)
    record.append(secondaryField('PID'))
    record.append(secondaryField('UR'))
    altUR = secondaryField('AltUR')
    if (d.scriptType == 'match') and (altUR in d.altURrec) and (d.altURrec[altUR] is not None) and (d.secondaryRecNo != d.altURrec[altUR]):
        altUR += '[dup]'
    record.append(altUR)
    if fileType == 1:
        record += ['', confidence]
    elif fileType == 3:
        record += ['', '']
    elif fileType == 4:
        record += ['', '', confidence]
        record += ['', '']
    for col in d.reportingColumns:
        if col in d.secondaryHas:
            if col in d.reportingDates  :
                thisCol = secondaryField(col)
                if (thisCol != '') and (thisCol is not None):
                    thisDate = d.sc.secondaryParseDate(thisCol)
                    if thisDate is not None:
                        record.append(thisDate)
                    else:
                        record.append(thisCol)
                else:
                    record.append(thisCol)
            else:
                record.append(secondaryField(col))
        else:
            record.append('')
    d.worksheet[thisFile].append(record)



def PrintMaster(checked, isChecked, fileType, masterRecNo, thisFile, message, confidence, i):
    '''
Print a master PMI record from a row in masterDetails
checked == True if this file that gets checked
fileType defines the number of UR/altUR fields to prepend to the heading
    1 - prepend a data columns of master PMI PID and UR (the default)
    2 - ONLY USED for secondary only files
    3 - prepend data columns for both of the above (secondary columns will be blank)
    4 - prepend heading columns for both of the above and the extensive match confidence levels (secondary columns and extensive match confidence levels will be blank)
    '''

    record = []
    if checked:
        record = [isChecked] + record
    record.append(message)
    record += ['', '', '']
    record.append(d.masterDetails[masterRecNo]['PID'])
    record.append(d.masterDetails[masterRecNo]['UR'])
    if fileType == 4:
        record += ['']
        if d.scriptType == 'match':
            record += [d.possExtensiveMatches[d.secondaryRecNo][confidence][i][1]]
            record += [d.possExtensiveMatches[d.secondaryRecNo][confidence][i][2]]
        else:
            record += [d.possExtensiveFinds[d.secondaryRecNo][confidence][i][1]]
            record += [d.possExtensiveFinds[d.secondaryRecNo][confidence][i][2]]
    for col in d.reportingColumns:
        if col in d.masterHas:
            if col in d.reportingDates  :
                if (d.masterDetails[masterRecNo][col] != '') and (d.masterDetails[masterRecNo][col] is not None):
                    thisDate = d.mc.masterParseDate(d.masterDetails[masterRecNo][col])
                    if thisDate is not None:
                        record.append(thisDate)
                    else:
                        record.append(d.masterDetails[masterRecNo][col])
                else:
                    record.append(d.masterDetails[masterRecNo][col])
            else:
                record.append(d.masterDetails[masterRecNo][col])
        else:
            record.append('')
    d.worksheet[thisFile].append(record)


def PrintNoMatch():
    '''
Print a record for a secondary records that was not matched with any master record
    '''

    PrintSecondary(False, '', 2, 'ud', f'{d.secondaryLongName} {d.secondaryAltURname} not found in {d.masterLongName}', 0)


def PrintMatchFound(masterRecNo):
    '''
Print a matched/found master record
    '''

    # We may have matched/found a merged patient, or an alias.
    # If the real patient has a different UR number to this merged patient or alias
    # then that's worth noting in a special report
    if masterRecNo in d.masterNewRec :    # Check merged
        newMasterRecNo = d.masterNewRec[masterRecNo]
        UR1 = d.masterDetails[masterRecNo]['UR']
        UR2 = d.masterDetails[newMasterRecNo]['UR']
        if UR1 != UR2:
            if d.scriptType == 'match':
                d.mmatch += 1
                PrintSecondary(False, '', 3, 'mm', '', 0)
                PrintMaster(False, '', 3, masterRecNo, 'mm', f'matched to a merged patent (with a different {d.masterURname})', 0, 0)
                PrintMaster(False, '', 3, newMasterRecNo, 'mm', 'who has been merged to', 0, 0)
                d.worksheet['mm'].append([''])
            else:
                d.mfound += 1
                PrintSecondary(False, '', 3, 'fm', '', 0)
                PrintMaster(False, '', 3, masterRecNo, 'fm', f'similar to a merged patient (with a different {d.masterURname})', 0, 0)
                PrintMaster(False, '', 3, newMasterRecNo, 'fm', 'who has been merged to', 0, 0)
                d.worksheet['fm'].append([''])
    elif masterRecNo in d.masterPrimRec :    # Check alias
        newMasterRecNo = d.masterPrimRec[masterRecNo]
        UR1 = d.masterDetails[masterRecNo]['UR']
        UR2 = d.masterDetails[newMasterRecNo]['UR']
        if UR1 != UR2:
            if d.scriptType == 'match':
                d.amatch += 1
                PrintSecondary(False, '', 3, 'ma', '', 0)
                PrintMaster(False, '', 3, masterRecNo, 'ma', f'matched to an alias (with a diferent {d.masterURname})', 0, 0)
                PrintMaster(False, '', 3, newMasterRecNo, 'ma', 'for this patient', 0, 0)
                d.worksheet['ma'].append([''])
            else:
                d.afound += 1
                PrintSecondary(False, '', 3, 'fa', '', 0)
                PrintMaster(False, '', 3, masterRecNo, 'fa', f'similar to an alias (with different {d.masterURname})', 0, 0)
                PrintMaster(False, '', 3, newMasterRecNo, 'fa', 'for this patient', 0, 0)
                d.worksheet['fa'].append([''])

    if d.scriptType == 'match':
        d.match += 1
        if not d.quick:
            if (d.match > 0) and ((d.match % 10000) == 0):
                if d.secondaryExtractDir:
                    PrintClose('m', 0, 6, f'{d.secondaryDir}/{d.secondaryExtractDir}/{d.secondaryShortName}_ur_matched_{d.matchVolume}.xlsx')
                else:
                    PrintClose('m', 0, 6, f'{d.secondaryDir}/{d.secondaryShortName}_ur_matched_{d.matchVolume}.xlsx')
                d.matchVolume += 1
                PrintHeading('m', 3, 'Message', '')
            PrintSecondary(False, '', 3, 'm', '', 0)
            PrintMaster(False, '', 3, masterRecNo, 'm', 'matched to', 0, 0)
            if masterRecNo in d.masterNewRec :            # Check merged
                newMasterRecNo = d.masterNewRec[masterRecNo]
                PrintMaster(False, '', 3, newMasterRecNo, 'm', 'who has been merged to', 0, 0)
            elif masterRecNo in d.masterPrimRec :        # Check alias
                newMasterRecNo = d.masterPrimRec[masterRecNo]
                PrintMaster(False, '', 3, newMasterRecNo, 'm', 'who is an alias for', 0, 0)
            d.worksheet['m'].append([''])
    else:
        isFound = CheckIfFound()
        check = ''
        if isFound:
            if (d.founddn > 0) and ((d.founddn % 10000) == 0):
                if d.secondaryExtractDir:
                    PrintClose('fdn', 0, 7, f'{d.secondaryDir}/{d.secondaryExtractDir}/{d.secondaryShortName}_Found_Done_{d.foundDoneVolume}.xlsx')
                else:
                    PrintClose('fdn', 0, 7, f'{d.secondaryDir}/{d.secondaryShortName}_Found_Done_{d.foundDoneVolume}.xlsx')
                d.foundDoneVolume += 1
                PrintHeading('fdn', 3, 'Checked', 'Message')
            d.founddn += 1
            check = 'Y'
            thisFile = 'fdn'
        else:
            if (d.foundtd > 0) and ((d.foundtd % 10000) == 0):
                if d.secondaryExtractDir:
                    PrintClose('ftd', 0, 7, f'{d.secondaryDir}/{d.secondaryExtractDir}/{d.secondaryShortName}_Found_ToDo_{d.foundToDoVolume}.xlsx')
                else:
                    PrintClose('ftd', 0, 7, f'{d.secondaryDir}/{d.secondaryShortName}_Found_ToDo_{d.foundToDoVolume}.xlsx')
                d.foundToDoVolume += 1
                PrintHeading('ftd', 3, 'Checked', 'Message')
            d.foundtd += 1
            check = ''
            thisFile = 'ftd'
        PrintSecondary(True, check, 3, thisFile, '', 0)
        PrintMaster(True, '', 3, masterRecNo, thisFile, 'found to be', 0, 0)
        if masterRecNo in d.masterNewRec :            # Check merged
            newMasterRecNo = d.masterNewRec[masterRecNo]
            PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who has been merged to', 0, 0)
        elif masterRecNo in d.masterPrimRec :        # Check alias
            newMasterRecNo = d.masterPrimRec[masterRecNo]
            PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who is an alias for', 0, 0)
        d.worksheet[thisFile].append([''])


def PrintMisMatch(masterRecNo, thisFile, message):
    '''Print Miss Mattched records'''

    # If there are extras display those as well
    misMessage = 'Mismatch on ' + message
    if d.extras[d.secondaryRecNo] != '':
        misMessage += ' <' + d.extras[d.secondaryRecNo] +'>'
    if masterRecNo in d.masterNewRec :        # Check merged
        newMasterRecNo = d.masterNewRec[masterRecNo]
        UR1 = d.masterDetails[masterRecNo]['UR']
        UR2 = d.masterDetails[newMasterRecNo]['UR']
        if UR1 != UR2:
            d.mmatch += 1
            newMasterRecNo = d.masterNewRec[masterRecNo]
            PrintSecondary(False, '', 3, 'mm', misMessage, 0)
            PrintMaster(False, '', 3, masterRecNo, 'mm', 'with patient', 0, 0)
            PrintMaster(False, '', 3, newMasterRecNo, 'mm', 'who has been merged to', 0, 0)
            d.worksheet['mm'].append([''])
            return
    elif masterRecNo in d.masterPrimRec :        # Check alias
        newMasterRecNo = d.masterPrimRec[masterRecNo]
        UR1 = d.masterDetails[masterRecNo]['UR']
        UR2 = d.masterDetails[newMasterRecNo]['UR']
        if UR1 != UR2:
            d.amatch += 1
            newMasterRecNo = d.masterPrimRec[masterRecNo]
            PrintSecondary(False, '', 3, 'ma', misMessage, 0)
            PrintMaster(False, '', 3, masterRecNo, 'ma', 'with patient', 0, 0)
            PrintMaster(False, '', 3, newMasterRecNo, 'ma', 'who is an alias for', 0, 0)
            d.worksheet['ma'].append([''])
            return

    # We need to print out mismatch for the secRecNo record in the secondary PMI
    isChecked = False
    check = ''
    pid = d.sc.secondaryCleanPID()
    if pid in d.matchedUR:
        isChecked = True
        if d.matchedUR[pid] == secondaryField('AltUR'):
            check = 'Y'
        else:
            check = f'ERROR in matched.xlsx: Incorrect {d.secondaryAltURname} number - {pid}, {d.matchedUR[pid]}'

    if message == 'Family Name [close]':
        d.fncmis += 1
        if isChecked:
            d.fncmisdn += 1
    elif message == 'Family Name Only':
        d.fnomis += 1
        if isChecked:
            d.fnomisdn += 1
    elif message == 'Family Name Plus':
        d.fnpmis += 1
        if isChecked:
            d.fnpmisdn += 1
    elif message == 'Birthdate':
        d.dobmis += 1
        if isChecked:
            d.dobmisdn += 1
    elif message == 'Given Names':
        d.gnmis += 1
        if isChecked:
            d.gnmisdn += 1
    elif message == 'Sex':
        d.sexmis += 1
        if isChecked:
            d.sexmisdn += 1
    if isChecked:
        PrintSecondary(True, check, 3, thisFile + 'dn', misMessage, 0)
    else:
        PrintSecondary(True, '', 3, thisFile + 'td', misMessage, 0)
    if isChecked:
        PrintMaster(True, '', 3, masterRecNo, thisFile + 'dn', '', 0, 0)
    else:
        PrintMaster(True, '', 3, masterRecNo, thisFile + 'td', '', 0, 0)
    if masterRecNo in d.masterNewRec :    # Check merged
        newMasterRecNo = d.masterNewRec[masterRecNo]
        if isChecked:
            PrintMaster(True, '', 3, newMasterRecNo, thisFile + 'dn', 'who has been merged to', 0, 0)
        else:
            PrintMaster(True, '', 3, newMasterRecNo, thisFile + 'td', 'who has been merged to', 0, 0)
    elif masterRecNo in d.masterPrimRec :    # Check alias
        newMasterRecNo = d.masterPrimRec[masterRecNo]
        if isChecked:
            PrintMaster(True, '', 3, newMasterRecNo, thisFile + 'dn', 'who is an alias for', 0, 0)
        else:
            PrintMaster(True, '', 3, newMasterRecNo, thisFile + 'td', 'who is an alias for', 0, 0)
    if isChecked:
        d.worksheet[thisFile + 'dn'].append([''])
    else:
        d.worksheet[thisFile + 'td'].append([''])


def PrintExtensiveMatch():
    '''
Print out the Possible matches
    '''

    masterRecNo = d.foundRec[d.secondaryRecNo]

    # We need to print out the extensive matches for the secondaryRecNo record in the secondary PMI
    thisFile = ''
    for confidence in (reversed(sorted(d.possExtensiveMatches[d.secondaryRecNo]))):
        if thisFile == '':
            if confidence == 100.0:
                extensiveMessage = 'extensive match'
                thisFile = 'em'
                d.extensivematch += 1
            else:
                extensiveMessage = 'probable match'
                thisFile = 'pm'
                d.probablematch += 1

        PrintSecondary(False, '', 4, thisFile, '', confidence)
        for i in range(len(d.possExtensiveMatches[d.secondaryRecNo][confidence])):
            masterRecNo = d.possExtensiveMatches[d.secondaryRecNo][confidence][i][0]
            PrintMaster(False, '', 4, masterRecNo, thisFile, extensiveMessage, confidence, i)
            if masterRecNo in d.masterNewRec :    # Check merged
                newMasterRecNo = d.masterNewRec[masterRecNo]
                PrintMaster(False, '', 4, newMasterRecNo, thisFile, 'who has been merged to', confidence, i)
            elif masterRecNo in d.masterPrimRec :    # Check alias
                newMasterRecNo = d.masterPrimRec[masterRecNo]
                PrintMaster(False, '', 4, newMasterRecNo, thisFile, 'who is an alias for', confidence, i)
        d.worksheet[thisFile].append([''])



def PrintDuplicateFound(message):
    '''
Print out Duplicate found master records for this secondary PMI record
    '''

    isFound = CheckIfFound()
    check = ''
    duplicateMessage =  ''
    if message == '':
        duplicateMessage = 'Duplicate Matches for this patient'
        if isFound:
            check = 'Y'
            thisFile = 'dfdn'
            d.dfounddn += 1
        else:
            check = ''
            thisFile = 'dftd'
            d.dfoundtd += 1
    else:
        if isFound:
            check = 'Y'
            thisFile = 'dpfdn'
            d.dfounddn += 1
        else:
            check = ''
            thisFile = 'dpftd'
            d.dpfoundtd += 1
        if message == 'Sound':
            duplicateMessage = 'Duplicate Similar sounding patients'
        else:
            duplicateMessage = 'Duplicate Similar patients - different ' + message

    PrintSecondary(True, check, 3, thisFile, duplicateMessage, 0)
    for i, masterRecNo in enumerate(d.URrecNos):
        masterSound = d.URrecSounds[i]
        if masterSound != '':
            masterSound = '[' + masterSound + ']'
        PrintMaster(True, '', 3, masterRecNo, thisFile, masterSound, 0, 0)
        if masterRecNo in d.masterNewRec :    # Check merged
            newMasterRecNo = d.masterNewRec[masterRecNo]
            PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who has been merged to' + masterSound, 0, 0)
        elif masterRecNo in d.masterPrimRec :    # Check alias
            newMasterRecNo = d.masterPrimRec[masterRecNo]
            PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who is an alias for' + masterSound, 0, 0)
    d.worksheet[thisFile].append([''])



def PrintPartialFound(thisFile, masterRecNo, message, masterRecSound):
    '''
Print out Similar found master records for this secondary PMI record
    '''

    isFound = CheckIfFound()
    check = ''
    if isFound:
        check = 'Y'
        thisFile += 'dn'
    else:
        check = ''
        thisFile += 'td'
    if message == 'Sound':
        duplicateMessage = 'Similar sounding patient'
    else:
        duplicateMessage = 'Similar patient - different ' + message
    PrintSecondary(True, check, 3, thisFile, duplicateMessage, 0)
    masterSound = masterRecSound
    if masterSound != '':
        masterSound = '[' + masterSound + ']'
    PrintMaster(True, '', 3, masterRecNo, thisFile, masterSound, 0, 0)
    if masterRecNo in d.masterNewRec :    # Check merged
        newMasterRecNo = d.masterNewRec[masterRecNo]
        PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who has been merged to' + masterSound, 0, 0)
    elif masterRecNo in d.masterPrimRec :    # Check alias
        newMasterRecNo = d.masterPrimRec[masterRecNo]
        PrintMaster(True, '', 3, newMasterRecNo, thisFile, 'who is an alias for' + masterSound, 0, 0)
    d.worksheet[thisFile].append([''])



def PrintClose(thisFile, fileType, reportCol, fileName):
    '''
Close a workbook
fileType is the number of columns before the d.reportingColumns columns
    '''

    heading = True
    cell = d.worksheet[thisFile]['A2']
    d.worksheet[thisFile].freeze_panes = cell
    for row in d.worksheet[thisFile].iter_rows():
        if heading:
            heading = False
            continue
        if fileType == 1:
            row[6].number_format = '0.00'
            row[7].number_format = '0.00'
            row[8].number_format = '0.00'
        if reportCol >= 0:
            for i, col in enumerate(d.reportingColumns):
                if col in d.reportingDates  :
                    if (row[reportCol + i].value != '') and (row[reportCol + i].value is not None):
                        row[reportCol + i].style = d.date_style        # Set the date style
    cleanUpExcel(d.worksheet[thisFile])
    try:
        d.workbook[thisFile].save(filename=fileName)
    except:
        logging.fatal('cannot create %s', fileName)
        sys.exit(EX_CANTCREAT)


def PrintExtensiveFinds():
    '''
Print out the Possible matches
    '''

    masterRecNo = d.foundRec[d.secondaryRecNo]

    # We need to print out the extensive matches for the secondaryRecNo record in the secondary PMI
    thisFile = ''
    for confidence in (reversed(sorted(d.possExtensiveFinds[d.secondaryRecNo]))):
        if thisFile == '':
            if confidence == 100.0:
                extensiveMessage = 'found similar patient'
                thisFile = 'ef'
                d.extensivefound += 1
            else:
                extensiveMessage = 'probable similar patient'
                thisFile = 'pf'
                d.probablefound += 1

        PrintSecondary(False, '', 4, thisFile, extensiveMessage, confidence)
        for i in range(len(d.possExtensiveFinds[d.secondaryRecNo][confidence])):
            masterRecNo = d.possExtensiveFinds[d.secondaryRecNo][confidence][i][0]
            PrintMaster(False, '', 4, masterRecNo, thisFile, '', confidence, i)
            if masterRecNo in d.masterNewRec :    # Check merged
                newMasterRecNo = d.masterNewRec[masterRecNo]
                PrintMaster(False, '', 4, newMasterRecNo, thisFile, 'who has been merged to', confidence, i)
            elif masterRecNo in d.masterPrimRec :    # Check alias
                newMasterRecNo = d.masterPrimRec[masterRecNo]
                PrintMaster(False, '', 4, newMasterRecNo, thisFile, 'who is an alias for', confidence, i)
        d.worksheet[thisFile].append([''])
