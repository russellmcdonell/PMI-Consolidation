
# pylint: disable=line-too-long

'''
A script to process the output files create by findUR.py

SYNOPSIS
$ python findpy secondaryDirectory [-c] [-f secondaryExtractDirectory|--secondaryExtractDir=secondaryExtractDirectory]
                                       [-v loggingLevel|--verbose=loggingLevel] [-o logfile|--logfile=logfile]


OPTIONS
secondaryDirectory
The directory containing the secondary configuration (secondary.cfg)
This directory may contain subdirectories where specific extracts will be found.
If an extract sub-directory is not specified, then this directory must hold all the report files created by findUR.py

-c|--csv
Create CSV files as well as the normal Excel file

-f secondaryExtractDir|--secondaryExtractDir=secondaryExtractDir
The optional extract sub-directory, of the secondary directory, containing the report files created by findUR.py

-v loggingLevel|--verbose=loggingLevel
Set the level of logging that you want.

-o logfile|--logfile=logfile
The name of a log file where you want all messages captured.


THE MAIN CODE
Start by parsing the command line arguements and setting up logging.

Then read in the findUR.py report files and create three files
 - found.xlsx being a list of the PIDs and UR pairs that were deemed a found match
 - notFound.xlsx being a list of the PIDs that were deemed to be not in the master PMI
 - patch.xlsx being a list of PID and UR pairs, where the AltUR is not the current UR for the patient.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

# Import the required modules
import sys
import os
import argparse
import logging
import glob
import re
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import data as d
import functions as f

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0        # successful termination
EX_WARN = 1        # non-fatal termination with warnings

EX_USAGE = 64        # command line usage error
EX_DATAERR = 65        # data format error
EX_NOINPUT = 66        # cannot open input
EX_NOUSER = 67        # addressee unknown
EX_NOHOST = 68        # host name unknown
EX_UNAVAILABLE = 69    # service unavailable
EX_SOFTWARE = 70    # internal software error
EX_OSERR = 71        # system error (e.g., can't fork)
EX_OSFILE = 72        # critical OS file missing
EX_CANTCREAT = 73    # can't create (user) output file
EX_IOERR = 74        # input/output error
EX_TEMPFAIL = 75    # temp failure; user is invited to retry
EX_PROTOCOL = 76    # remote error in protocol
EX_NOPERM = 77        # permission denied
EX_CONFIG = 78        # configuration error


if __name__ == '__main__':
    '''
The main code
Start by parsing the command line arguements and setting up logging.
Then read in the findUR.py report files and create
found.xlsx - PID,UR
notFound.xlsx - PID
patch.xlsx - AltUR,UR
    '''

    # Save the program name
    d.progName = sys.argv[0]
    d.progName = d.progName[0:-3]        # Strip off the .py ending
    d.scriptType = 'found'

    # Get the options
    parser = argparse.ArgumentParser(description='Check the goodness of health of a secondary PMI extraction')
    parser.add_argument ('secondaryDir', metavar='secondaryDirectory', help='The name of directory containg the secondary configuration and cleanSecondary.py routines')
    parser.add_argument ('-c', '--csv', dest='wantCSV', action='store_true', help='Create CSV files as well as the normal Excel file')
    parser.add_argument ('-f', '--secondaryExtractDir', dest='secondaryExtractDir', metavar='secondaryExtractDirectory', default=None, help='The name of the secondary directory sub-directory that contains the extract secondary CSV file and configuration specific to the extract')
    parser.add_argument ('-v', '--verbose', dest='verbose', type=int, choices=range(0,5), help='The level of logging\n\t0=CRITICAL,1=ERROR,2=WARNING,3=INFO,4=DEBUG')
    parser.add_argument ('-o', '--logfile', dest='logfile', metavar='logfile', default=None, help='The name of a logging file')
    args = parser.parse_args()

    logging_levels = {0:logging.CRITICAL, 1:logging.ERROR, 2:logging.WARNING, 3:logging.INFO, 4:logging.DEBUG}
    logfmt = d.progName + ' [%(asctime)s]: %(message)s'
    if args.verbose:    # Change the logging level from "WARN" if the -v vebose option is specified
        loggingLevel = args.verbose
        if args.logfile :        # and send it to a file if the -o logfile option is specified
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel], filename=args.logfile)
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel])
    else:
        if args.logfile :        # send the default (WARN) logging to a file if the -o logfile option is specified
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', filename=args.logfile)
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p')


    d.secondaryDir = args.secondaryDir
    wantCSV = args.wantCSV
    d.secondaryExtractDir = args.secondaryExtractDir

    # Read in the secondary configuration file
    f.getSecondaryConfig(False)

    # Read in the extract configuration file if required
    if d.secondaryExtractDir:
        # Read in the extract configuration file
        f.getSecondaryConfig(True)

    # Open Error file
    f.openErrorFile()

    # Open the output files
    foundWB = Workbook()
    foundWS = foundWB.active
    foundWS.append(["PID","UR"])
    foundCount = 0
    found = set()
    notFoundWB = Workbook()
    notFoundWS = notFoundWB.active
    notFoundWS.append(["PID"])
    notFoundCount = 0
    notFound = set()
    patchWB = Workbook()
    patchWS = patchWB.active
    patchWS.append(["PID","UR"])
    toPatchCount = 0
    patch = set()
    toDoCount = 0
    if wantCSV:
        if d.secondaryExtractDir:
            filenameFound = f'{d.secondaryDir}/{d.secondaryExtractDir}/found.csv'
        else:
            filenameFound = f'{d.secondaryDir}/found.csv'
        try:
            csvFileFound = open(filenameFound, 'wt', newline='')
            csvWriterFound = csv.writer(csvFileFound, dialect=csv.excel)
            csvWriterFound.writerow(['PID', 'UR'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/found.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/found.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)
        if d.secondaryExtractDir:
            filenameNotFound = f'{d.secondaryDir}/{d.secondaryExtractDir}/notFound.csv'
        else:
            filenameNotFound = f'{d.secondaryDir}/notFound.csv'
        try:
            csvFileNotFound = open(filenameNotFound, 'wt', newline='')
            csvWriterNotFound = csv.writer(csvFileNotFound, dialect=csv.excel)
            csvWriterNotFound.writerow(['PID'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/notFound.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/notFound.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)
        if d.secondaryExtractDir:
            filenamePatch = f'{d.secondaryDir}/{d.secondaryExtractDir}/patch.csv'
        else:
            filenamePatch = f'{d.secondaryDir}/patch.csv'
        try:
            csvFilePatch = open(filenamePatch, 'wt', newline='')
            csvWriterPatch = csv.writer(csvFilePatch, dialect=csv.excel)
            csvWriterPatch.writerow(['PID', 'UR'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/patch.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/patch.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)

    # Read in the Found PIDs - matches found and no checking required (d.found)
    f.getFound()
    for pid, ur in d.foundUR.items():
        thisFound = pid + '~' + ur
        if thisFound in found:
            logging.info('Skipping %s,%s - in found.xlsx', pid, ur)
            continue
        foundWS.append([pid, ur])
        found.add(thisFound)
        foundCount += 1
        if wantCSV:
            csvWriterFound.writerow([pid, ur])

    # Read in the Not Found PIDs - confirmed non-matches - further checking requried (d.notFound)
    f.getNotFound()
    for pid in d.notFound:
        if pid in notFound:
            logging.info('Skipping %s - in notFound.xlsx', pid)
            continue
        notFoundWS.append([pid])
        notFound.add(pid)
        notFoundCount += 1
        if wantCSV:
            csvWriterNotFound.writerow([pid])

    # Read in any PID/UR number that need to be patched
    patched = {}
    haveWorkbook = True
    try:
        if d.secondaryExtractDir:
            toPatchWB = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/patch.xlsx')
        else:
            toPatchWB = load_workbook(f'./{d.secondaryDir}/patch.xlsx')
        toPatchWS = toPatchWB.active
    except:
        haveWorkbook = False

    if haveWorkbook:
        heading = True
        for row in toPatchWS.iter_rows():
            if heading:
                heading = False
                continue
            if len(row) != 2:
                if d.secondaryExtractDir:
                    logging.fatal('Input error in ./%s/%s/patch.xlsx - wrong number of fields', d.secondaryDir, d.secondaryExtractDir)
                else:
                    logging.fatal('Input error in ./%s/patch.xlsx - wrong number of fields', d.secondaryDir)
                sys.exit(EX_DATAERR)
            PID = str(row[0].value)
            UR = str(row[1].value)
            thisPatch = PID + '~' + UR
            if thisPatch in patch:
                logging.info('Skipping %s,%s - in patch.xlsx', PID, UR)
                continue
            patchWS.append([PID, UR])
            patch.add(thisPatch)
            toPatchCount += 1
            if wantCSV:
                csvWriterPatch.writerow([PID, UR])


    # Read in and process all the "_SimilarFound*_ToDo.xlsx"
    # One secondary record and one master record
    if d.secondaryExtractDir:
        foundFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_SimilarFound_*_ToDo.xlsx')
    else:
        foundFiles = glob.glob(f'{d.secondaryDir}/*_SimilarFound_ToDo.xlsx')
    for file in foundFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        thisPID = None
        canMatch = canFix = False
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            # "n" means the PID is not found
            checked = row[0].value
            if checked is None:
                checked = ''
            elif not isinstance(checked, str):
                logging.fatal('Illegal entry in "Checked" column (%s) - aborting', str(row[0].value))
                sys.exit(EX_DATAERR)
            if (row[1].value is not None) and isinstance(row[1].value, str) and (row[1].value[0:7] == 'Similar') and (row[2].value is not None):      # A secondary record
                thisPID = row[2].value
                if checked.lower() == 'n':
                    if thisPID in notFound:
                        logging.info('Skipping not finding %s - in %s', thisPID, os.path.basename(file))
                        canMatch = canFix = False
                        continue
                    notFoundWS.append([thisPID])
                    notFound.add(thisPID)
                    if wantCSV:
                        csvWriterNotFound.writerow([thisPID])
                    notFoundCount += 1
                    canMatch = isMatched = canFix = False
                elif checked.lower() == 'y':
                    canMatch = True
                    canFix = False
                elif checked == '':
                    toDoCount += 1
                    logging.warning('%s: Not checked, PID(%s)', os.path.basename(file), thisPID)
                    canMatch = canFix = False
                else:
                    logging.fatal('Illegal entry in "Checked" column (%s) in file %s - aborting', str(checked), os.path.basename(file))
            elif (row[5].value is not None) and (row[6].value is not None):      # The master record
                if canMatch:
                    thisFound = thisPID + '~' + row[6].value
                    if thisFound in found:
                        logging.info('Skipping finding %s,%s - in %s', thisPID, row[6].value, os.path.basename(file))
                        canMatch = False
                        canFix = True
                        continue
                    foundWS.append([thisPID, row[6].value])
                    found.add(thisFound)
                    if wantCSV:
                        csvWriterFound.writerow([thisPID, row[6].value])
                    foundCount += 1
                    canMatch = False
                    canFix = True
                elif canFix and ((checked == 'who has been merged to') or (checked == 'who is an alias for')) :      # A master record
                    thisPatch = thisPID + '~' + row[6].value
                    if thisPatch in patch:
                        logging.info('Skipping patching %s,%s - in %s', thisPID, row[6].value, os.path.basename(file))
                        continue
                    patchWS.append([thisPID, row[6].value])
                    patch.add(thisPatch)
                    if wantCSV:
                        csvWriterPatch.writerow([thisPID, row[6].value])
                    toPatchCount += 1
                thisPID = None
                canMatch = canFix = False
        logging.info('%d toDo, %d found, %d notFound after file %s', toDoCount, foundCount, notFoundCount, os.path.basename(file))

    # Read in and process all the "_Duplicate*Found*.xlsx"
    # One secondary record and multiple master records
    if d.secondaryExtractDir:
        foundFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_Duplicate*Found_ToDo.xlsx')
    else:
        foundFiles = glob.glob(f'{d.secondaryDir}/*_Duplicate*Found_ToDo.xlsx')
    for file in foundFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        thisPID = thisUR = thisMerge = None
        canMatch = isMatched = canFix = False
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            # "n" means the PID is not found
            checked = row[0].value
            if checked is None:
                checked = ''
            elif not isinstance(checked, str):
                logging.fatal('Illegal entry in "Checked" column (%s) - aborting', str(checked))
                sys.exit(EX_DATAERR)
            if (row[1].value is not None) and isinstance(row[1].value, str) and (row[1].value[0:9] == 'Duplicate') and (row[2].value is not None):      # A secondary record
                thisPID = row[2].value
                if checked.lower() == 'n':
                    if thisPID in notFound:
                        logging.info('Skipping not finding %s - in %s', thisPID, os.path.basename(file))
                        canMatch = isMatched = canFix = False
                        continue
                    notFoundWS.append([thisPID])
                    notFound.add(thisPID)
                    if wantCSV:
                        csvWriterNotFound.writerow([thisPID])
                    notFoundCount += 1
                    canMatch = isMatched = canFix = False
                elif checked.lower() == 'y':
                    canMatch = True
                    isMatched = canFix = False
                elif checked == '':
                    toDoCount += 1
                    logging.warning('%s: Not checked, PID(%s)', os.path.basename(file), thisPID)
                    thisPID = thisUR = thisMerge = None
                    canMatch = canFix = False
                else:
                    logging.fatal('Illegal entry in "Checked" column (%s) in file %s - aborting', str(checked), os.path.basename(file))
                    sys.exit(EX_DATAERR)
            elif (row[5].value is not None) and (row[6].value is not None):      # A master record
                if checked.lower() != 'y':
                    continue
                if canMatch:                # We have seen a 'y' on a secondary record
                    thisUR = row[6].value
                    isMatched = canFix = True
                    canMatch = False
                elif isMatched:
                    logging.fatal('More than one "Y" for secondary record (%s), URs(%s and %s) in file %s - aborting', thisPID, thisUR, row[5].value, file)
                    sys.exit(EX_DATAERR)
                elif canFix and (checked == 'who has been merged to') or (checked == 'who is an alias for') :      # A master record
                    thisMerge = row[6].value
            else:           # The blank line
                if canMatch:
                    logging.fatal('No master record "Y" for secondary record (%s) in file %s - aborting', thisPID, file)
                    sys.exit(EX_DATAERR)
                elif isMatched:
                    thisFound = thisPID + '~' + thisUR
                    if thisFound in found:
                        logging.info('Skipping finding %s,%s - in %s', thisPID, thisUR, os.path.basename(file))
                        continue
                    foundWS.append([thisPID, thisUR])
                    found.add(thisFound)
                    if wantCSV:
                        csvWriterFound.writerow([thisPID, thisUR])
                    foundCount += 1
                    if thisMerge is not None:
                        thisPatch = thisPID + '~' + thisMerge
                        if thisPatch in patch:
                            logging.info('Skipping patching %s,%s - in %s', thisPID, thisMerge, os.path.basename(file))
                            continue
                        patchWS.append([thisPID, thisMerge])
                        if wantCSV:
                            csvWriterPatch.writerow([thisPID, thisMerge])
                        toPatchCount += 1
                thisPID = thisUR = thisMerge = None
                canMatch = isMatched = canFix = False
        if canMatch:            # Process the last record(s) - there may not be a blank line to end the file
            logging.fatal('No master record "Y" for secondary record (%s) in file %s - aborting', thisPID, file)
            sys.exit(EX_DATAERR)
        elif isMatched:
            thisFound = thisPID + '~' + thisUR
            if thisFound in found:
                logging.info('Skipping finding %s,%s - in %s', thisPID, thisUR, os.path.basename(file))
                continue
            foundWS.append([thisPID, thisUR])
            found.add(thisFound)
            if wantCSV:
                csvWriterFound.writerow([thisPID, thisUR])
            foundCount += 1
            if thisMerge is not None:
                thisPatch = thisPID + '~' + thisMerge
                if thisPatch in patch:
                    logging.info('Skipping patching %s,%s - in %s', thisPID, thisMerge, os.path.basename(file))
                    continue
                patchWS.append([thisPID, thisMerge])
                if wantCSV:
                    csvWriterPatch.writerow([thisPID, thisMerge])
                toPatchCount += 1
        logging.info('%d ToDo, %d found, %d notFound after file %s', toDoCount, foundCount, notFoundCount, os.path.basename(file))

    # Read in and process all the "_Found_ToDo_n.xlsx"
    # One secondary record and one master record
    if d.secondaryExtractDir:
        foundFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_Found_ToDo_[1-9]*.xlsx')
    else:
        foundFiles = glob.glob(f'{d.secondaryDir}/*_Found_ToDo_[1-9]*.xlsx')
    for file in foundFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        thisPID = thisUR = None
        canMatch = False
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            # "n" means the PID is not found
            checked = row[0].value
            if checked is None:
                checked = ''
            elif not isinstance(checked, str):
                logging.fatal('Illegal entry in "Checked" column (%s) - aborting', str(checked))
                sys.exit(EX_DATAERR)
            if row[2].value is not None:      # A secondary record
                thisPID = row[2].value
                if checked.lower() == 'n':
                    if thisPID in notFound:
                        logging.info('Skipping not finding %s - in %s', thisPID, os.path.basename(file))
                        continue
                    notFoundWS.append([thisPID])
                    notFound.add(thisPID)
                    if wantCSV:
                        csvWriterNotFound.writerow([thisPID])
                    notFoundCount += 1
                    thisPID = thisUR = thisMerge = None
                    canMatch = canFix = False
                elif checked.lower() == 'y':
                    canMatch = True
                elif checked == '':
                    toDoCount += 1
                    logging.warning('%s: Not checked, PID(%s)', os.path.basename(file), thisPID)
                    thisPID = thisUR = thisMerge = None
                    canMatch = canFix = False
                else:
                    logging.fatal('Illegal entry in "Checked" column (%s) - aborting', str(checked))
                    sys.exit(EX_DATAERR)
            elif (row[5].value is not None) and (row[6].value is not None):      # the master record
                if canMatch:                # We have seen a 'y' on a secondary record
                    thisUR = row[6].value
                    thisFound = thisPID + '~' + thisUR
                    if thisFound in found:
                        logging.info('Skipping finding %s,%s - in %s', thisPID, thisUR, os.path.basename(file))
                        continue
                    foundWS.append([thisPID, thisUR])
                    found.add(thisFound)
                    if wantCSV:
                        csvWriterFound.writerow([thisPID, thisUR])
                    foundCount += 1
                    thisPID = thisUR = thisMerge = None
                    canMatch = canFix = False
            else:           # The blank line
                thisPID = thisUR = thisMerge = None
                canMatch = canFix = False
        logging.info('%d ToDo, %d found, %d notFound after file %s', toDoCount, foundCount, notFoundCount, os.path.basename(file))

    # Read in and process the "_NotFound_ToDo.xlsx" file
    # One secondary record
    if d.secondaryExtractDir:
        foundFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_NotFound_ToDo.xlsx')
    else:
        foundFiles = glob.glob(f'{d.secondaryDir}/*_NotFound_ToDo.xlsx')
    for file in foundFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            # "AltUR" means the PID is found
            checked = row[0].value
            thisPID = row[2].value
            if checked is None:
                checked = ''
            if checked == '':
                if thisPID in notFound:
                    logging.info('Skipping not finding %s - in %s', thisPID, os.path.basename(file))
                else:
                    notFoundWS.append([thisPID])
                    notFound.add(thisPID)
                    if wantCSV:
                        csvWriterNotFound.writerow([thisPID])
                    notFoundCount += 1
                continue
            thisUR = checked
            thisFound = thisPID + '~' + thisUR
            if thisFound in found:
                logging.info('Skipping finding %s,%s - in %s', thisPID, thisUR, os.path.basename(file))
                continue
            foundWS.append([thisPID, thisUR])
            found.add(thisFound)
            if wantCSV:
                csvWriterFound.writerow([thisPID, thisUR])
            foundCount += 1
        logging.info('%d ToDo, %d found, %d notFound after file %s', toDoCount, foundCount, notFoundCount, os.path.basename(file))

    # Save the output files
    if d.secondaryExtractDir:
        foundWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/found.xlsx')
        notFoundWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/notFound.xlsx')
        patchWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/patch.xlsx')
    else:
        foundWB.save(filename=f'{d.secondaryDir}/found.xlsx')
        notFoundWB.save(filename=f'{d.secondaryDir}/notFound.xlsx')
        patchWB.save(filename=f'{d.secondaryDir}/patch.xlsx')
    if wantCSV:
        csvFileFound.close()
        csvFileNotFound.close()
        csvFilePatch.close()

    # And finally, create the report
    f.openReport()
    heading = 'Found/not found Processing - found'
    d.rpt.write(f'{heading}\n')
    heading = re.sub('[^ ]', '_', heading)
    d.rpt.write(f'{heading}\n')
    d.rpt.write(f'{foundCount}\tmatches found (found.xlsx)\n')
    d.rpt.write(f'{notFoundCount}\twith no matching master PMI record (notFound.xlsx)\n')
    d.rpt.write(f'{toPatchCount}\tmatches found, but to aliases or merged patients (patch.xlsx)\n')
    d.rpt.write(f'{toDoCount}\trecords still to be checked')

    d.rpt.close()

    # Close the error log csv file and exit
    d.fe.close()
    sys.exit(EX_OK)
