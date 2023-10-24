'''Clean a master extract file'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, line-too-long, unspecified-encoding, anomalous-backslash-in-string

import sys
import csv
import logging
import re
import datetime
import unicodedata
from openpyxl import load_workbook
from  dateutil.parser import parse
import functions as f
import data as d

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0        # successful termination
EX_WARN = 1        # non-fatal termination with warnings

EX_USAGE = 64        # command line usage error
EX_DATAERR = 65        # data format error
EX_NOINPUT = 66        # cannot open input
EX_NOUSER = 67        # addressee unknown
EX_NOHOST = 68        # host name unknown
EX_UNAVAILABLE = 69    # service unavailable
EX_SOFTWARE = 70    # internal software error
EX_OSERR = 71        # system error (e.g., can't fork)
EX_OSFILE = 72        # critical OS file missing
EX_CANTCREAT = 73    # can't create (user) output file
EX_IOERR = 74        # input/output error
EX_TEMPFAIL = 75    # temp failure; user is invited to retry
EX_PROTOCOL = 76    # remote error in protocol
EX_NOPERM = 77        # permission denied
EX_CONFIG = 78        # configuration error

def masterOpenRawPMI():
    '''
Open a raw master PMI extract.
Set up a csv parser for parsing it if appropriate.
    '''

    d.masterIsCSV = False

    try:
        if d.masterIsCSV:
            d.mfr = open(d.masterFileName, 'rt')        # For CSV files
        else:
            d.mwb = load_workbook(d.masterFileName)  # For Excel files
    except:
        logging.fatal('cannot open raw master PMI file (%s)', d.masterFileName)
        sys.exit(EX_NOINPUT)
    if d.masterIsCSV:
        # For CSV files
        sample = d.mfr.read(4096)
        d.mfr.seek(0)
        d.dialect = csv.Sniffer().sniff(sample)
        d.dialect.skipinitialspace = True
        d.dialect.doublequote = True
        d.masterHasHeader = csv.Sniffer().has_header(sample)
    else:
        # For Excel files
        d.mws = d.mwb.active
        # d.mws = d.mwb['Master PMI']
        d.mws_iter_rows = d.mws.iter_rows()
        d.masterHasHeader = 0
    return


def masterCloseRawPMI():
    '''
Close the master raw PMI file
    '''
    if d.masterIsCSV:
        d.mfr.close()       # For CSV files
    return


def masterReadRawPMI():
    '''
Read the next record of the raw PMI extract file.
    '''

    while True :        # Keep reading lines until we have something to return
        try:
            if d.masterIsCSV:
                line = d.mfr.readline()     # CSV file
            else:
                line = next(d.mws_iter_rows)    # Excel file
        except StopIteration:
            return False
        except:
            return False

        if d.masterIsCSV:
            # For CSV files
            if line == '':
                return False
            line = line.strip()

            # Skip the heading of the master PMI extra file has one
            if d.masterHasHeader:
                d.masterHasHeader = False
                continue

            # Clean up the line if necessary
            nfkd_form = unicodedata.normalize('NFKD', line)
            line = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

            # Split into CSV fields
            for row in csv.reader([line], d.dialect):
                csvifields = row
                break
        else:
            # For Excel files
            if d.masterHasHeader < 1:            # Skip header line
                d.masterHasHeader += 1
                continue
            csvifields = []
            if line[0].value is None:               # Skipp blank lines (usually at end of worksheet)
                continue
            for cell in line:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        text = cell.value
                        nfkd_form = unicodedata.normalize('NFKD', text)
                        cleanText = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])
                        csvifields.append(cleanText)
                    else:
                        csvifields.append(str(cell.value))
                else:
                    csvifields.append('')
            if len(csvifields) > d.masterFieldCount:     # Ignore extra columns
                csvifields = csvifields[0:d.masterFieldCount]

        # Flag one more raw record read in
        d.masterRawRecNo += 1

        # Check if correct number of columns
        if len(csvifields) != int(d.masterFieldCount):
            d.feCSV.writerow([f'INPUT FIELD COUNT ERROR [found:{len(csvifields)}, expected {d.masterFieldCount}], record No:{d.masterRawRecNo}'])
            d.feCSV.writerow(csvifields)
            continue

        # Save the required columns
        d.csvfields = []
        for i, col in enumerate(d.masterSaveColumns):
            if col == -1:
                d.csvfields.append(d.masterRawRecNo)
            elif col < -1 :        # Save a filler to be fixed up later
                d.csvfields.append('')
            else:
                value = csvifields[int(col)]

                # Fix up some unprintable characters
                value = re.sub('\xE1', '', value)

                # Save cleaned up value
                d.csvfields.append(value)

        # Now save any derived columns
        for i, col in enumerate(d.masterSaveColumns):
            if col < -1 :         # Save a derived PID
                # value = d.mc.masterField('UR') + '^' + d.mc.masterField('Alias')
                value = ''
                d.csvfields[i] = value

        # Check if deleted record
        if d.ml.masterIsDeleted():
            d.dCount += 1
            continue

        # Check validity of the PID
        pid = masterCleanPID()
        if pid == '':
            d.feCSV.writerow(['MISSING PID NUMBER'])
            d.feCSV.writerow(d.csvfields)
            continue

        # Some tests associated with numeric PID values - comment out if PID values are not number
        '''
        intPID = f.masterIntPID()

        # If the PID is a number, then intPID should be equal to pid, with any leading zeros stripped
        if intPID != re.sub('^0*', '', pid):
            d.feCSV.writerow([f'NON NUMERIC {d.masterPIDname} ({pid}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        '''

        # Check the validity of the UR value
        '''
        ur = masterCleanUR()
        if ur == '':
            d.feCSV.writerow([f'MISSING {d.masterURname} in record ({d.masterRawRecNo}) being discarded'])
            continue
        if len(ur) > 7:
            d.feCSV.writerow([f'ILLEGALLY LONG {d.masterURname} ({ur}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if len(ur) < 7:
            d.feCSV.writerow([f'ILLEGALLY SHORT {d.masterURname} ({ur}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        '''

        # Some tests associated with numeric UR values - comment out if UR values are not number
        intUR = f.masterIntUR()

        # If the UR is a number, then intUR should be equal to ur, with any leading zeros stripped
        '''
        if intUR != re.sub('^0*', '', ur):
            d.feCSV.writerow([f'NON NUMERIC {d.masterURname} ({ur}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        # If the UR starts with '9' then it's not in a valid range
        if ur[0] == '9':
            d.feCSV.writerow([f'{d.masterURname} from a foreing range ({ur}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        '''

        # Check if UR has been seen already
        if intUR in d.URasIntRec:
            d.feCSV.writerow([f'Ambiguos {d.masterURname} ({intUR}) in record ({d.masterRawRecNo})', 'records Nos:', d.URasIntRec[intUR], d.masterRawRecNo])
        else:
            d.URasIntRec[intUR] = d.masterRecNo

        # Some tests for family names and given names that imply that the current record is not a valid record
        # Checks for family names that imply skip this record
        '''
        familyName = f.masterField('FamilyName')
        givenName = f.masterField('GivenName')
        if re.search('^REF*ER [TUR]', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow([f'REFER TO PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if re.search('^Use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow([f'"Use" PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if re.search('^Do not use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow([f'"Do not use" PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if re.search('^Dont use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow([f'"Do not use" PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if re.search(r'^C\/E', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow([f'CORRECTION ERROR PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        if (re.search('unknown', familyName, flags=re.IGNORECASE) is not None) and (re.search(r'^C\/E', givenName, flags=re.IGNORECASE) is not None):
            d.feCSV.writerow(['UNKNOWN,C/E PATIENT ({familyName}) in record ({d.masterRawRecNo}) being discarded'])
            continue
        '''

        return True
    return False


def masterCleanUR():
    '''
Clean up a UR value
    '''

    ur = f.masterField('UR')
    '''
    ur = re.sub(' *$', '', ur)    # Strip of any trailing blanks
    '''
    return ur


def masterCleanPID():
    '''
Clean up a PID value
    '''

    pid = f.masterField('PID')
    '''
    pid = re.sub(' *$', '', pid)    # Strip of any trailing blanks
    '''
    return pid


def masterNeatFamilyName():
    '''
Neaten up the family name by cleaning up thing users may have 'added'.
    '''

    familyName = f.masterField('FamilyName')
    familyName = familyName.upper().strip()

    familyName = re.sub(r' *\(.*\)', '', familyName)        # Things in round brackets
    familyName = re.sub(r' *<.*>', '', familyName)            # Things in angle brackets
    familyName = re.sub(r' *[\'"`].*[\'"`]', '', familyName)        # Things in quotes
    familyName = re.sub(r'^\*', '', familyName)            # A leading asterisk
    familyName = re.sub(r'\.', '', familyName)            # Full stops
    familyName = re.sub(r'\**$', '', familyName)            # A training asterisks
    familyName = re.sub(r'\'', '', familyName)            # Single quotes

    return familyName


def masterCleanFamilyName():
    '''
Clean up the family name by removing things that make matching difficult
    '''

    familyName = d.mc.masterNeatFamilyName()                # Start with a neat family name
    familyName = re.sub(r'~', '', familyName)               # This is mandatory

    familyName = re.sub(r'`', '', familyName)               # Remove backquotes
    familyName = re.sub(r',', '', familyName)               # Remove commas
    familyName = re.sub(r'  *', ' ', familyName)            # Remove redundant extra spaces

    return familyName


def masterNeatGivenName():
    '''
Neaten up the given name by cleaning up thing users may have 'added'.
    '''

    givenName = f.masterField('GivenName')
    givenName = givenName.upper().strip()
    '''
    givenName = re.sub(r' *\(.*\)', '', givenName)        # Things in round brackets
    givenName = re.sub(r' *[\'"`].*[\'"`]', '', givenName)    # Things in quotes
    givenName = re.sub(r' *[\'"`(][A-Z:]*$', '', givenName)    # Things after quotes
    givenName = re.sub(r',$', '', givenName)            # Trailing commas
    givenName = re.sub(r'\/[A-Z\/]*$', '', givenName)    # Names after slashes
    givenName = re.sub(r'\.', '', givenName)            # Remove full stops
    givenName = re.sub(r'\**$', '', givenName)        # A training asterisks
    '''
    return givenName


def masterCleanGivenName():
    '''
Clean up the given name by removing things that make matching difficult
    '''

    givenName = d.mc.masterNeatGivenName()             # Start with a neat given name
    givenName = re.sub(r'~', '', givenName)            # This is mandatory
    '''
    givenName = re.sub(r'`', '', givenName)            # Remove backquotes
    givenName = re.sub(r',', '', givenName)            # Remove commas
    givenName = re.sub(r'  *', ' ', givenName)          # Remove redundant extra spaces
    '''
    return givenName


def masterCleanDOB():
    '''
Clean up a Birth date
This routine handles mismatches between the master PMI extra file format of a date of birth and the required internal format [ISO  8601:'YYYY-MM-DD']
    '''

    # Get the Birthdate from the master file
    dob = f.masterField('Birthdate')
    dob = re.sub(r'~', '', dob)    # This is mandatory

    # Raw format is d[d]/m[m]/yyyy
    '''
    dob = re.sub(r' .*', '', dob)    # Remove any potiential time value
    bits = re.split(r'\/', dob)    # Split into day, month, year
    if len(bits) != 3 :        # Check for potentially invalid date
        return ''
    try:
        bDay = int(bits[0])
        bMonth = int(bits[1])
        bYear = int(bits[2])
    except:
        return ''
    '''

    # Raw format is ISO 8601:YYYY-MM-DD
    dob = re.sub('T.*', '', dob)    # Remove any potiential time value
    bits = re.split(r'-', dob)    # Split into day, month, year
    if len(bits) != 3 :        # Check for potentially invalid date
        return ''
    try:
        bYear = int(bits[0])
        bMonth = int(bits[1])
        bDay = int(bits[2])
    except:
        return ''

    # Raw format is ISO 8601:YYYYMMDD
    '''
    if len(dob) < 8:
        return ''
    try:
        bYear = int(dob[0:4])
        bMonth = int(dob[4:6])
        bDay = int(dob[6:8])
    except:
        return ''
    '''

    # Convert to an internal data value
    try:
        thisDate = datetime.date(bYear, bMonth, bDay)    # Check for valid date
    except:
        return ''

    # Check that this is not a future date
    if thisDate > d.today:
        return ''

    # Return the required internal format [ISO 8601:YYYY-MM-DD]
    return thisDate.isoformat()



def masterParseDate(thisDate):
    '''
Parse a date string - we assume all date strings in the master file are formatted the same
    '''

    try:
        # Raw format is d[d]/m[m]/yyyy
        # masterDate = parse(thisDate, dayfirst=True).date()

        # Raw format is ISO 8601:YYYY-MM-DD
        masterDate = parse(thisDate, dayfirst=False).date()

        return masterDate
    except:
        return None



def masterCleanSex():
    '''
Clean up the Sex value
This routine handles mismatches between the master PMI extra file format for sex and the internal required format of M/F
    '''

    # Raw format is Male, Female, male, female
    sex = f.masterField('Sex')
    sex = re.sub('~', '', sex)    # This is mandatory

    # Return the required format
    if sex != '':
        sex = sex[0].upper()
    return sex
