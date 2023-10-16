
# pylint: disable=line-too-long

'''
A script to process the output files create by matchAltUR.py

SYNOPSIS
$ python matched.py secondaryDirectory [-c] [-f secondaryExtractDirectory|--secondaryExtractDir=secondaryExtractDirectory]
                                       [-v loggingLevel|--verbose=loggingLevel] [-o logfile|--logfile=logfile]


OPTIONS
secondaryDirectory
The directory containing the secondary configuration (secondary.cfg)
This directory may contain subdirectories where specific extracts will be found.
If an extract sub-directory is not specified, then this directory must hold all the report files created by matchAltUR.py

-c|--csv
Create CSV files as well as the normal Excel file

-f secondaryExtractDir|--secondaryExtractDir=secondaryExtractDir
The optional extract sub-directory, of the secondary directory, containing the report files created by matchAltUR.py

-v loggingLevel|--verbose=loggingLevel
Set the level of logging that you want.

-o logfile|--logfile=logfile
The name of a log file where you want all messages captured.


THE MAIN CODE
Start by parsing the command line arguements and setting up logging.

Then read in the matchAltUR.py report files and create three files
 - matched.xlsx being a list of the PIDs that were deemed a match
 - notMatched.xlsx being a list of the PIDs that were deemed to be not a match
 - fix.xlsx being a list of AltUR and UR pairs, where the AltUR is not the current UR for the patient.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

# Import the required modules
import sys
import os
import argparse
import logging
import glob
import re
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import data as d
import functions as f

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0        # successful termination
EX_WARN = 1        # non-fatal termination with warnings

EX_USAGE = 64        # command line usage error
EX_DATAERR = 65        # data format error
EX_NOINPUT = 66        # cannot open input
EX_NOUSER = 67        # addressee unknown
EX_NOHOST = 68        # host name unknown
EX_UNAVAILABLE = 69    # service unavailable
EX_SOFTWARE = 70    # internal software error
EX_OSERR = 71        # system error (e.g., can't fork)
EX_OSFILE = 72        # critical OS file missing
EX_CANTCREAT = 73    # can't create (user) output file
EX_IOERR = 74        # input/output error
EX_TEMPFAIL = 75    # temp failure; user is invited to retry
EX_PROTOCOL = 76    # remote error in protocol
EX_NOPERM = 77        # permission denied
EX_CONFIG = 78        # configuration error


if __name__ == '__main__':
    '''
The main code
Start by parsing the command line arguements and setting up logging.
Then read in the matchAltUR.py report files and create
matched.xlsx - PID,UR
notMatched.xlsx - PID
fix.xlsx - AltUR,UR
    '''

    # Save the program name
    d.progName = sys.argv[0]
    d.progName = d.progName[0:-3]        # Strip off the .py ending
    d.scriptType = 'matched'

    # Get the options
    parser = argparse.ArgumentParser(description='Check the goodness of health of a secondary PMI extraction')
    parser.add_argument ('secondaryDir', metavar='secondaryDirectory', help='The name of directory containg the secondary configuration and cleanSecondary.py routines')
    parser.add_argument ('-c', '--csv', dest='wantCSV', action='store_true', help='Create CSV files as well as the normal Excel file')
    parser.add_argument ('-f', '--secondaryExtractDir', dest='secondaryExtractDir', metavar='secondaryExtractDirectory', default=None, help='The name of the secondary directory sub-directory that contains the extract secondary CSV file and configuration specific to the extract')
    parser.add_argument ('-v', '--verbose', dest='verbose', type=int, choices=range(0,5), help='The level of logging\n\t0=CRITICAL,1=ERROR,2=WARNING,3=INFO,4=DEBUG')
    parser.add_argument ('-o', '--logfile', dest='logfile', metavar='logfile', default=None, help='The name of a logging file')
    args = parser.parse_args()

    logging_levels = {0:logging.CRITICAL, 1:logging.ERROR, 2:logging.WARNING, 3:logging.INFO, 4:logging.DEBUG}
    logfmt = d.progName + ' [%(asctime)s]: %(message)s'
    if args.verbose:    # Change the logging level from "WARN" if the -v vebose option is specified
        loggingLevel = args.verbose
        if args.logfile :        # and send it to a file if the -o logfile option is specified
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel], filename=args.logfile)
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel])
    else:
        if args.logfile :        # send the default (WARN) logging to a file if the -o logfile option is specified
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', filename=args.logfile)
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p')


    d.secondaryDir = args.secondaryDir
    wantCSV = args.wantCSV
    d.secondaryExtractDir = args.secondaryExtractDir

    # Read in the secondary configuration file
    f.getSecondaryConfig(False)

    # Read in the extract configuration file if required
    if d.secondaryExtractDir:
        # Read in the extract configuration file
        f.getSecondaryConfig(True)

    # Open Error file
    f.openErrorFile()

    # Open the output files
    matchedWB = Workbook()
    matchedWS = matchedWB.active
    matchedWS.append(["PID","UR"])
    matchedCount = 0
    matched = set()
    notMatchedWB = Workbook()
    notMatchedWS = notMatchedWB.active
    notMatchedWS.append(["PID"])
    notMatchedCount = 0
    notMatched = set()
    fixWB = Workbook()
    fixWS = fixWB.active
    fixWS.append(["AltUR","UR"])
    toFixCount = 0
    fix = set()
    toDoCount = 0
    if wantCSV:
        if d.secondaryExtractDir:
            filenameMatched = f'{d.secondaryDir}/{d.secondaryExtractDir}/matched.csv'
        else:
            filenameMatched = f'{d.secondaryDir}/matched.csv'
        try:
            csvFileMatched = open(filenameMatched, 'wt', newline='')
            csvWriterMatched = csv.writer(csvFileMatched, dialect=csv.excel)
            csvWriterMatched.writerow(['PID', 'UR'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/matched.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/matched.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)
        if d.secondaryExtractDir:
            filenameNotMatched = f'{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.csv'
        else:
            filenameNotMatched = f'{d.secondaryDir}/notMatched.csv'
        try:
            csvFileNotMatched = open(filenameNotMatched, 'wt', newline='')
            csvWriterNotMatched = csv.writer(csvFileNotMatched, dialect=csv.excel)
            csvWriterNotMatched.writerow(['PID'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/notMatched.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/notMatched.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)
        if d.secondaryExtractDir:
            filenameFix = f'{d.secondaryDir}/{d.secondaryExtractDir}/fix.csv'
        else:
            filenameFix = f'{d.secondaryDir}/fix.csv'
        try:
            csvFileFix = open(filenameFix, 'wt', newline='')
            csvWriterFix = csv.writer(csvFileFix, dialect=csv.excel)
            csvWriterFix.writerow(['AltUR', 'UR'])
        except:
            if d.secondaryExtractDir:
                logging.fatal('Cannot create ./%s/%s/fix.csv', d.secondaryDir, d.secondaryExtractDir)
            else:
                logging.fatal('Cannot create ./%s/fix.csv', d.secondaryDir)
            sys.exit(EX_CANTCREAT)

    # Read in the Matched PIDs - matches found and no checking required (d.matched)
    f.getMatched()
    for pid, ur in d.matchedUR.items():
        thisMatch = pid + '~' + ur
        if thisMatch in matched:
            logging.info('Skipping %s,%s - in matched.xlsx', pid, ur)
            continue
        matchedWS.append([pid, ur])
        matched.add(thisMatch)
        matchedCount += 1
        if wantCSV:
            csvWriterMatched.writerow([pid, ur])

    # Read in the Not Matched PIDs - confirmed non-matches - further checking requried (d.notMatched)
    f.getNotMatched()
    for pid in d.notMatched:
        if pid in notMatched:
            logging.info('Skipping %s - in notMatched.xlsx', pid)
            continue
        notMatchedWS.append([pid])
        notMatched.add(pid)
        notMatchedCount += 1
        if wantCSV:
            csvWriterNotMatched.writerow([pid])

    # Read in any AltUR number that need to be fixed
    fixed = {}
    haveWorkbook = True
    try:
        if d.secondaryExtractDir:
            toFixWB = load_workbook(f'./{d.secondaryDir}/{d.secondaryExtractDir}/fix.xlsx')
        else:
            toFixWB = load_workbook(f'./{d.secondaryDir}/fix.xlsx')
        toFixWS = toFixWB.active
    except:
        haveWorkbook = False

    if haveWorkbook:
        heading = True
        for row in toFixWS.iter_rows():
            if heading:
                heading = False
                continue
            if len(row) != 2:
                if d.secondaryExtractDir:
                    logging.fatal('Input error in ./%s/%s/fix.xlsx - wrong number of fields', d.secondaryDir, d.secondaryExtractDir)
                else:
                    logging.fatal('Input error in ./%s/fix.xlsx - wrong number of fields', d.secondaryDir)
                sys.exit(EX_DATAERR)
            AltUR = str(row[0].value)
            UR = str(row[1].value)
            thisFix = AltUR + '~' + UR
            if thisFix in fix:
                logging.info('Skipping %s,%s - in fix.xlsx', AltUR, UR)
                continue
            fixWS.append([AltUR, UR])
            fix.add(thisFix)
            toFixCount += 1
            if wantCSV:
                csvWriterFix.writerow([AltUR, UR])

    # Read in and process all the "_ur_matched_n.xlsx"
    # Here we assume 'Y' unless there is an 'N'
    if d.secondaryExtractDir:
        matchedFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_ur_matched_[1-9]*.xlsx')
    else:
        matchedFiles = glob.glob(f'{d.secondaryDir}/*_ur_matched_[1-9]*.xlsx')
    for file in matchedFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        canFix = False
        thisPID = thisAltUR = None
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            # "n" means the PID is not matched
            checked = row[0].value
            if (checked is None) or (not isinstance(checked, str)):
                checked = ''
            if (row[1].value is not None) and (row[3].value is not None):      # A secondary record
                thisPID = row[1].value
                thisAltUR = row[3].value
                if thisAltUR[-5:] == '[dup]':
                    thisAltUR = thisAltUR[:-5]
                if checked.lower() == 'n':
                    if thisPID in notMatched:
                        logging.info('Skipping %s - in %s', thisPID, file)
                        continue
                    notMatchedWS.append([thisPID])
                    notMatched.add(thisPID)
                    if wantCSV:
                        csvWriterNotMatched.writerow([thisPID])
                    notMatchedCount += 1
                    thisAltUR = None
                    canFix = False
                else:
                    canFix = True
            elif (row[4].value is not None) and (row[5].value is not None):      # A master record
                if canFix and (checked == 'matched to'):
                    thisMatch = thisPID + '~' + thisAltUR
                    if thisMatch in matched:
                        logging.info('Skipping %s,%s - in %s', thisPID, thisAltUR, file)
                        continue
                    matchedWS.append([thisPID, thisAltUR])
                    matched.add(thisMatch)
                    if wantCSV:
                        csvWriterMatched.writerow([thisPID, thisAltUR])
                    matchedCount += 1
                elif canFix and ((checked == 'who has been merged to') or (checked == 'who is an alias for')) :      # A master record
                    thisFix = thisAltUR + '~' + row[5].value
                    if thisFix in fix:
                        logging.info('Skipping %s,%s - in %s', thisAltUR, row[5].value, file)
                        continue
                    fixWS.append([thisAltUR, row[5].value])
                    fix.add(thisFix)
                    if wantCSV:
                        csvWriterFix.writerow([thisAltUR, row[5].value])
                    toFixCount += 1
                else:
                    thisAltUR = None
                    canFix = False
            else:
                thisAltUR = None
                canFix = False

    # Read in and process all the "_mismatch_ToDo.xlsx" files
    # Here a 'Y' means a match, a 'N' means not a match and nothing means 'yet to be done'
    if d.secondaryExtractDir:
        matchedFiles = glob.glob(f'{d.secondaryDir}/{d.secondaryExtractDir}/*_mismatch_ToDo.xlsx')
    else:
        matchedFiles = glob.glob(f'{d.secondaryDir}/*_mismatch_ToDo.xlsx')
    for file in matchedFiles:
        try:
            wb = load_workbook(file)
            ws = wb.active
        except:
            logging.fatal('Error reading %s', file)
            sys.exit(EX_IOERR)

        heading = True
        canFix = False
        thisPID = thisAltUR = None
        for row in ws.iter_rows():
            if heading:
                heading = False
                continue
            checked = row[0].value
            if (checked is None) or (not isinstance(checked, str)):
                checked = ''
            message = row[1].value
            if (message is None) or (not isinstance(message, str)):
                message = ''
            if (message[0:12] == 'Mismatch on ') and (row[2].value is not None) and (row[4].value is not None):      # A secondary record
                thisPID = row[2].value
                thisAltUR = row[4].value
                # "n" means the PID is not matched
                if checked.lower() == 'n':
                    if thisPID in notMatched:
                        logging.info('Skipping %s - in %s', thisPID, os.path.basename(file))
                        continue
                    notMatchedWS.append([thisPID])
                    notMatched.add(thisPID)
                    if wantCSV:
                        csvWriterNotMatched.writerow([thisPID])
                    notMatchedCount += 1
                    canFix = False
                # "y" one a secondary row means the PID is matched to the next master UR
                elif checked.lower() == 'y':
                    thisMatch = thisPID + '~' + thisAltUR
                    if thisMatch in matched:
                        logging.info('Skipping %s,%s - in %s', thisPID, thisAltUR, os.path.basename(file))
                        continue
                    matchedWS.append([thisPID, thisAltUR])
                    matched.add(thisMatch)
                    if wantCSV:
                        csvWriterMatched.writerow([thisPID, thisAltUR])
                    matchedCount += 1
                    canFix = True
                else:               # neither "y" nor "n" - ignore, but log alias/merges
                    toDoCount += 1
                    if checked != '':
                        logging.warning('%s: illegal character (%s) in checked, PID(%s), AltUR(%s)', checked, os.path.basename(file), thisPID, thisAltUR)
                    canFix = True
            elif (row[5].value is not None) and (row[6].value is not None):      # A master record
                if canFix and ((message == 'who has been merged to') or (message == 'who is an alias for')):
                    thisFix = thisAltUR + '~' + row[6].value
                    if thisFix in fix:
                        logging.info('Skipping %s,%s - in %s', thisAltUR, row[6].value, file)
                        continue
                    fixWS.append([thisAltUR, row[6].value])
                    fix.add(thisFix)
                    if wantCSV:
                        csvWriterFix.writerow([thisAltUR, row[6].value])
                    toFixCount += 1
                    canFix = False
            else:
                canFix = False

    # Read in and process the "_ur_undefined.xlsx" file
    # Here we assume 'not matched, unless there is a 'Y'
    if d.secondaryExtractDir:
        filename =  f'{d.secondaryDir}/{d.secondaryExtractDir}/{d.secondaryShortName}_ur_undefined.xlsx'
    else:
        filename =  f'{d.secondaryDir}/{d.secondaryShortName}_ur_undefined.xlsx'
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except:
        logging.fatal('Error reading %s', filename)
        sys.exit(EX_IOERR)

    heading = True
    canFix = False
    thisPID = thisAltUR = None
    for row in ws.iter_rows():
        if heading:
            heading = False
            continue
        # "y" one a secondary row means the PID is matched to the AltUR
        checked = row[0].value
        thisPID = row[1].value
        if (checked is None) or (not isinstance(checked, str)):
            checked = ''
        if (checked.lower() == 'y') and (row[1].value is not None) and (row[3].value is not None):      # A secondary record
            thisAltUR = row[3].value
            if thisAltUR[-5:] == '[dup]':
                thisAltUR = thisAltUR[:-5]
            thisMatch = thisPID + '~' + thisAltUR
            if thisMatch in matched:
                logging.info('Skipping %s,%s - in %s', thisPID, thisAltUR, os.path.basename(filename))
                continue
            matchedWS.append([thisPID, thisAltUR])
            matched.add(thisMatch)
            if wantCSV:
                csvWriterMatched.writerow([thisPID, thisAltUR])
            matchedCount += 1
        else:
            if thisPID in notMatched:
                logging.info('Skipping %s - in %s', thisPID, os.path.basename(filename))
                continue
            notMatchedWS.append([thisPID])
            notMatched.add(thisPID)
            if wantCSV:
                csvWriterNotMatched.writerow([thisPID])
            notMatchedCount += 1

    # Save the output files
    if d.secondaryExtractDir:
        matchedWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/matched.xlsx')
        notMatchedWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/notMatched.xlsx')
        fixWB.save(filename=f'{d.secondaryDir}/{d.secondaryExtractDir}/fix.xlsx')
    else:
        matchedWB.save(filename=f'{d.secondaryDir}/matched.xlsx')
        notMatchedWB.save(filename=f'{d.secondaryDir}/notMatched.xlsx')
        fixWB.save(filename=f'{d.secondaryDir}/fix.xlsx')
    if wantCSV:
        csvFileMatched.close()
        csvFileNotMatched.close()
        csvFileFix.close()

    # And finally, create the report
    f.openReport()
    heading = 'Matched/not Matched Processing - matched'
    d.rpt.write(f'{heading}\n')
    heading = re.sub('[^ ]', '_', heading)
    d.rpt.write(f'{heading}\n')
    d.rpt.write(f'{matchedCount}\tmatches (matched.xlsx)\n')
    d.rpt.write(f'{notMatchedCount}\tmismatches (notMatched.xlsx)\n')
    d.rpt.write(f'{toFixCount}\tmatches to aliases or merged patients (fix.xlsx)\n')
    d.rpt.write(f'{toDoCount}\trecords still to be checked')

    d.rpt.close()

    # Close the error log csv file and exit
    d.fe.close()
    sys.exit(EX_OK)
