'''Clean a secondary extract file'''

# pylint: disable=invalid-name, bare-except, unspecified-encoding, pointless-string-statement, line-too-long


import sys
import csv
import logging
import re
import datetime
import unicodedata
from openpyxl import load_workbook
from  dateutil.parser import parse
import functions as f
import data as d

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0        # successful termination
EX_WARN = 1        # non-fatal termination with warnings

EX_USAGE = 64        # command line usage error
EX_DATAERR = 65        # data format error
EX_NOINPUT = 66        # cannot open input
EX_NOUSER = 67        # addressee unknown
EX_NOHOST = 68        # host name unknown
EX_UNAVAILABLE = 69    # service unavailable
EX_SOFTWARE = 70    # internal software error
EX_OSERR = 71        # system error (e.g., can't fork)
EX_OSFILE = 72        # critical OS file missing
EX_CANTCREAT = 73    # can't create (user) output file
EX_IOERR = 74        # input/output error
EX_TEMPFAIL = 75    # temp failure; user is invited to retry
EX_PROTOCOL = 76    # remote error in protocol
EX_NOPERM = 77        # permission denied
EX_CONFIG = 78        # configuration error


def secondaryOpenRawPMI():
    '''
Open a raw secondary PMI extract.
Set up a csv parser for parsing it if appropriate.
    '''

    d.secondaryIsCSV = False

    try:
        if d.secondaryIsCSV:
            d.sfr = open(d.secondaryFileName, 'rt')     # For CSV files
        else:
            d.swb = load_workbook(d.secondaryFileName)  # For Excel files
    except:
        logging.fatal('cannot open raw secondary PMI file (%s)', d.secondaryFileName)
        sys.exit(EX_NOINPUT)

    if d.secondaryIsCSV:
        # For CSV files
        sample = d.sfr.read(4096)
        d.sfr.seek(0)
        d.dialect = csv.Sniffer().sniff(sample)
        d.dialect.skipinitialspace = False
        d.dialect.doublequote = True
        d.secondaryHasHeader = csv.Sniffer().has_header(sample)
    else:
        # For Excel files
        d.sws = d.swb.active
        # d.sws = d.swb['Secondary PMI']
        d.sws_iter_rows = d.sws.iter_rows()
        d.secondaryHasHeader = 0

    return


def secondaryCloseRawPMI():
    '''
Close the secondary raw PMI file
    '''
    if d.secondaryIsCSV:
        d.sfr.close()       # For CSV files
    return


def secondaryReadRawPMI():
    '''
Read the next record of the raw PMI extract file.
    '''

    while True :        # Keep reading lines until we have something to return
        try:
            if d.secondaryIsCSV:
                line = d.sfr.readline()         # CSV file
            else:
                line = next(d.sws_iter_rows)    # Excel file
        except StopIteration:
            return False
        except:
            return False
        if d.secondaryIsCSV:
            # For CSV files
            if line == '':
                return False
            line = line.rstrip()

            # Skip the heading of the secondary PMI extra file has one
            if d.secondaryHasHeader:
                d.secondaryHasHeader = False
                continue

            # Clean up the line if necessary
            nfkd_form = unicodedata.normalize('NFKD', line)
            line = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

            # Split into CSV fields
            for row in csv.reader([line], d.dialect):
                csvifields = row
                break
        else:
            # For Excel files
            if d.secondaryHasHeader < 3:            # Skip multpile header lines
                d.secondaryHasHeader += 1
                continue
            csvifields = []
            if line[0].value is None:               # Skipp blank lines (usually at end of worksheet)
                continue
            for cell in line:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        text = cell.value
                        nfkd_form = unicodedata.normalize('NFKD', text)
                        cleanText = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])
                        csvifields.append(cleanText)
                    else:
                        csvifields.append(str(cell.value))
                else:
                    csvifields.append('')
            if len(csvifields) > d.secondaryFieldCount:     # Ignore extra columns
                csvifields = csvifields[0:d.secondaryFieldCount]

        # Flag one more raw record read in
        d.secondaryRawRecNo += 1

        # Check if correct number of columns
        if len(csvifields) != int(d.secondaryFieldCount):
            d.feCSV.writerow([f'INPUT FIELD COUNT ERROR [found:{len(csvifields)}, expected {d.secondaryFieldCount}], record No:{d.secondaryRawRecNo}'])
            d.feCSV.writerow(csvifields)
            continue

        # Save the required columns
        d.csvfields = []
        for col in d.secondarySaveColumns:
            if col == -1:
                d.csvfields.append(str(d.secondaryRawRecNo))
            elif col < -1 :        # Save a filler to be fixed up later
                d.csvfields.append('')
            else:
                value = csvifields[int(col)]

                # Fix up some unprintable characters
                value = re.sub('\xE1', '', value)

                # Save cleaned up value
                d.csvfields.append(value)

        # Now save any derived columns
        for i, col in enumerate(d.secondarySaveColumns):
            if col < -1 :         # Save a derived PID
                # value = f.secondaryField('UR') + '^' + f.secondaryField('Alias')
                value = ''
                d.csvfields[i] = value

        # Check if deleted record
        if d.sl.secondaryIsDeleted():
            d.dCount += 1
            continue

        # Check validity of the PID
        pid = secondaryCleanPID()
        if pid == '':
            d.feCSV.writerow([f'MISSING {d.secondaryPIDname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue

        # Some tests associated with numeric PID values - comment out if PID values are not number
        '''
        intPID = f.secondaryIntPID()

        # If the PID is a number, then intPID should be equal to pid, with any leading zeros stripped
        if intPID != re.sub('^0*', '', pid):
            d.feCSV.writerow([f'NON NUMERIC {d.secondaryPIDname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue
        '''

        # Check the validity of the UR value
        ur = secondaryCleanUR()
        if ur == '':
            d.feCSV.writerow([f'MISSING {d.secondaryURname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue

        '''
        if len(ur) > 9:
            d.feCSV.writerow([f'ILLEGALLY LONG {d.secondaryURname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue
        if len(ur) < 9:
            d.feCSV.writerow(['fILLEGALLY SHORT {d.secondaryURname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue
        '''

        # Some tests associated with numeric UR values - comment out if UR values are not number
        '''
        intUR = f.secondaryIntUR()

        # If the UR is a number, then intUR should be equal to ur, with any leading zeros stripped
        if intUR != re.sub('^0*', '', ur):
            d.feCSV.writerow([f'NON NUMERIC {d.secondaryURname} - discarding record'])
            d.feCSV.writerow(d.csvfields)
            continue
        if intUR in d.URasIntRec:
            d.feCSV.writerow([f'Ambiguos {d.secondaryURname}: {intUR}', 'records Nos:', d.URasIntRec[intUR], d.secondaryRecNo])
        else:
            d.URasIntRec[intUR] = d.secondaryRecNo
        '''

        # Check the validity of the Alt UR value
        thisAltUR = secondaryCleanAltUR()                # Get the altUR number
        # intAltUR = f.secondaryIntAltUR()                # Get the intAltUR number

        if thisAltUR == '' :                        # Check if blank
            d.feCSV.writerow([f'Blank/Missing {d.secondaryAltURname} in record ({d.secondaryRawRecNo}) being ignored'])
            d.noAltURcount += 1
        else :                                # Check if valid
            '''
            if intAltUR != re.sub('^0*', '', thisAltUR):
                d.feCSV.writerow([f'NON NUMERIC {d.secondaryAltURname} ({thisAltUR}) in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            '''
            if len(thisAltUR) > 6:                # Check it's not too long
                d.feCSV.writerow([f'Illegally long {d.secondaryAltURname} ({thisAltUR}) in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            if len(thisAltUR) < 6:                # Check it's not too short
                d.feCSV.writerow([f'Illegally short {d.secondaryAltURname} ({thisAltUR}) in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            '''
            if thisAltUR[0:1] == '9':                # Check it starts with the digit 9
                d.feCSV.writerow([f'Illegal {d.secondaryAltURname} ({thisAltUR}) starting with the digit 9 in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            if intAltUR[0:1] == '-' :                    # Check for negative altUR numbers
                d.feCSV.writerow([f'Negative {d.secondaryAltURname} ({thisAltUR}) in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            if int(intAltUR) == 0 :                    # Check for zero altUR numbers
                d.feCSV.writerow([f'Zero {d.secondaryAltURname} ({thisAltUR}) in record ({d.secondaryRawRecNo}) being ignored'])
                thisAltUR = ''
            '''
            if thisAltUR == '':
                d.dudAltURcount += 1
                f.secondaryFieldSave('AltUR', '')    # Save a blank altUR

        # Some tests for family names and given names that imply that the current record is not a valid record
        # Checks for family names that imply skip this record
        familyName = f.secondaryField('FamilyName')
        givenName = f.secondaryField('GivenName')
        '''
        if re.search('^REF*ER [TUR]', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow(['REFER TO PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        if re.search('^Use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow(['"Use" PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        if re.search('^Do not use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow(['"Do not use" PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        if re.search('^Dont use ', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow(['"Do not use" PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        if re.search(r'^C\/E', familyName, flags=re.IGNORECASE) is not None:
            d.feCSV.writerow(['CORRECTION ERROR PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        '''
        r'''
        if (re.search('unknown', familyName, flags=re.IGNORECASE) is not None) and (re.search(r'^C\/E', givenName, flags=re.IGNORECASE) is not None):
            d.feCSV.writerow(['UNKNOWN,C/E PATIENT'])
            d.feCSV.writerow(d.csvfields)
            continue
        '''

        return True
    return False


def secondaryCleanUR():
    '''
Clean up a UR value
    '''

    ur = f.secondaryField('UR')
    '''
    ur = re.sub(' *$', '', ur)    # Strip of any trailing blanks
    '''
    return ur


def secondaryCleanAltUR():
    '''
Clean up a AltUR value
    '''

    altUR = f.secondaryField('AltUR')
    # altUR = re.sub(' *$', '', altUR)    # Strip of any trailing blanks
    return altUR


def secondaryCleanPID():
    '''
Clean up a PID value
    '''

    pid = f.secondaryField('PID')
    pid = re.sub(' *$', '', pid)    # Strip of any trailing blanks
    return pid


def secondaryNeatFamilyName():
    '''
Neaten up the family name by cleaning up thing users may have 'added'.
    '''

    familyName = f.secondaryField('FamilyName')
    familyName = familyName.upper().strip()
    familyName = re.sub(r' *\(.*\)', '', familyName)            # Things in round brackets
    familyName = re.sub(r' *<.*>', '', familyName)            # Things in angle brackets
    familyName = re.sub(r' *[\'"`].*[\'"`]', '', familyName)        # Things in quotes
    familyName = re.sub(r'^\*', '', familyName)            # A leading asterisk
    familyName = re.sub(r'\.', '', familyName)            # Full stops
    familyName = re.sub(r'\**$', '', familyName)            # A training asterisks
    familyName = re.sub(r'\'', '', familyName)            # Single quotes
    return familyName


def secondaryCleanFamilyName():
    '''
Clean up the family name by removing things that make matching difficult
    '''

    familyName = secondaryNeatFamilyName()                # Start with a neat family name
    familyName = re.sub(r'~', '', familyName)             # This is mandatory
    familyName = re.sub(r'`', '', familyName)             # Remove backquotes
    familyName = re.sub(r',', '', familyName)             # Remove commas
    familyName = re.sub(r'  *', ' ', familyName)          # Remove redundant extra spaces
    return familyName


def secondaryNeatGivenName():
    '''
Neaten up the given name by cleaning up thing users may have 'added'.
    '''

    givenName = f.secondaryField('GivenName')
    givenName = givenName.upper().strip()
    givenName = re.sub(r' *\(.*\)', '', givenName)        # Things in round brackets
    givenName = re.sub(r' *[\'"`].*[\'"`]', '', givenName)    # Things in quotes
    givenName = re.sub(r' *[\'"`(][A-Z:]*$', '', givenName)    # Things after quotes
    givenName = re.sub(r',$', '', givenName)            # Trailing commas
    givenName = re.sub(r'\/[A-Z\/]*$', '', givenName)    # Names after slashes
    givenName = re.sub(r'\.', '', givenName)            # Remove full stops
    givenName = re.sub(r'\**$', '', givenName)        # A training asterisks
    return givenName


def secondaryCleanGivenName():
    '''
Clean up the given name by removing things that make matching difficult
    '''

    givenName = secondaryNeatGivenName()            # Start with a neat given name
    givenName = re.sub(r'~', '', givenName)            # This is mandatory
    givenName = re.sub(r'`', '', givenName)            # Remove backquotes
    givenName = re.sub(r',', '', givenName)            # Remove commas
    givenName = re.sub(r'  *', ' ', givenName)         # Remove redundant extra spaces
    return givenName


def secondaryCleanDOB():
    '''
Clean up a Birth date
This routine handles mismatches between the secondary PMI extra file format of a date of birth and the required internal format [ISO  8601:'YYYY-MM-DD']
    '''

    # Get the Birthdate from the secondary file
    dob = f.secondaryField('Birthdate')
    dob = re.sub(r'~', '', dob)    # This is mandatory

    # Raw format is d[d]/m[m]/yyyy
    '''
    dob = re.sub(r' .*', '', dob)    # Remove any potiential time value
    bits = re.split(r'\/', dob)    # Split into day, month, year
    if len(bits) != 3 :        # Check for potentially invalid date
        return ''
    try:
        bDay = int(bits[0])
        bMonth = int(bits[1])
        bYear = int(bits[2])
    except:
        return ''
    '''

    # Raw format is ISO 8601:YYYY-MM-DD
    dob = re.sub('T.*', '', dob)    # Remove any potiential time value
    bits = re.split(r'-', dob)    # Split into day, month, year
    if len(bits) != 3 :        # Check for potentially invalid date
        return ''
    try:
        bYear = int(bits[0])
        bMonth = int(bits[1])
        bDay = int(bits[2])
    except:
        return ''

    # Raw format is ISO 8601:YYYYMMDD
    '''
    if len(dob) < 8:
        return ''
    try:
        bYear = int(dob[0:4])
        bMonth = int(dob[4:6])
        bDay = int(dob[6:8])
    except:
        return ''
    '''

    # Convert to an internal data value
    try:
        thisDate = datetime.date(bYear, bMonth, bDay)    # Check for valid date
    except:
        return ''

    # Check that this is not a future date
    if thisDate > d.today:
        return ''

    # Return the required internal format [ISO 8601:YYYY-MM-DD]
    return thisDate.isoformat()


def secondaryParseDate(thisDate):
    '''
Parse a date string - we assume all date strings in the master file are formatted the same
    '''

    try:
        # Raw format is d[d]/m[m]/yyyy
        secondaryDate = parse(thisDate, dayfirst=True).date()

        # Raw format is ISO 8601:YYYY-MM-DD
        # secondaryDate = parse(thisDate, dayfirst=False).date()

        return secondaryDate
    except:
        return None


def secondaryCleanSex():
    '''
Clean up the Sex value
This routine handles mismatches between the secondary PMI extra file format for sex and the internal required format of M/F
    '''

    # Raw format is Male, Female, male, female
    sex = f.secondaryField('Sex')
    sex = re.sub('~', '', sex)    # This is mandatory

    # Return the required format
    if sex != '':
        sex = sex[0].upper()
    return sex
